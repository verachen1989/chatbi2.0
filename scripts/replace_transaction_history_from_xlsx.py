#!/usr/bin/env python3
"""Replace dashboard transaction history with one authoritative detail workbook."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATA_RE = re.compile(
    r"const DATA = (.*?);\n(const PROJECT_METADATA_OVERRIDES|const LAUNCH_OFFICIAL_INVENTORY_OVERRIDES|const DEFAULT_PERIODS)",
    re.S,
)
PROJECT_ALIASES_RE = re.compile(r"const PROJECT_NAME_ALIASES = \{(.*?)\};", re.S)
OFFICIAL_NEW_LAUNCH_RE = re.compile(r"const ZJW_OFFICIAL_NEW_LAUNCH_PROJECTS = (\[.*?\]);", re.S)
PROJECT_BASIC_INFO_RE = re.compile(r"window\.PROJECT_BASIC_INFO\s*=\s*(\{.*?\});", re.S)
ALLOWED_PROPERTY_TYPES = {"普通住宅", "别墅"}
REPLACEMENT_START = (2025, 7)
PROJECT_FIELDS = (
    "project",
    "matchedName",
    "summaryRecordName",
    "officialProjectName",
    "cricProjectName",
    "janAprMatchedName",
    "junMatchedName",
    "junCricProjectName",
    "julMatchedName",
    "julCricProjectName",
)


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).lower()
    replacements = {
        "雲": "云",
        "澐": "云",
        "樹": "树",
        "華": "华",
        "鳴": "鸣",
        "號": "号",
        "鄕": "乡",
        "臺": "台",
        "灣": "湾",
        "萬": "万",
        "叁": "三",
        "贰": "二",
        "壹": "一",
        "玖": "九",
        "·": "",
        "•": "",
        ".": "",
        "。": "",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[\s\-—_/\\()（）【】\[\]《》“”\"'：:；;，,]+", "", text)


def split_names(value: Any) -> list[str]:
    return [part.strip() for part in re.split(r"[；;、/\n]+", clean(value)) if part.strip()]


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def rounded(value: float, digits: int = 2) -> int | float:
    result = round(float(value), digits)
    return int(round(result)) if abs(result - round(result)) < 0.000001 else result


def trade_amount_wan(value: Any) -> int | float:
    """Convert the CRIC detail amount from yuan to ten-thousand yuan."""
    return rounded(number(value) / 10000, 4)


def month_label(value: Any) -> str:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return f"{value.year % 100:02d}年{value.month}月"
    match = re.match(r"(\d{4})[-/](\d{1,2})", clean(value))
    if not match:
        return ""
    return f"{int(match.group(1)) % 100:02d}年{int(match.group(2))}月"


def month_tuple(label: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d{2})年(\d{1,2})月", label)
    return (2000 + int(match.group(1)), int(match.group(2))) if match else (0, 0)


def is_replacement_month(label: str) -> bool:
    return month_tuple(label) >= REPLACEMENT_START


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    return clean(value).split(" ", 1)[0].replace("-", "/")


def historical_house_key(row: dict[str, Any]) -> tuple[str, ...]:
    building = clean(row.get("building_name"))
    room = clean(row.get("room_number"))
    if building and room:
        return (
            clean(row.get("project_name")),
            clean(row.get("pre_permit")),
            building,
            clean(row.get("unit_number")),
            room,
        )
    return ("md5", clean(row.get("md5_str")) or repr(sorted(row.items())))


def summarize_rows(rows: list[dict[str, Any]]) -> dict[str, int | float]:
    area = sum(number(row.get("area")) for row in rows)
    amount = sum(number(row.get("totalWan")) for row in rows)
    return {
        "suites": len(rows),
        "area": rounded(area, 2),
        "amountWan": rounded(amount, 4),
        "avgPrice": int(round(amount * 10000 / area)) if area else 0,
    }


def parse_alias_groups(html: str) -> list[list[str]]:
    match = PROJECT_ALIASES_RE.search(html)
    if not match:
        return []
    groups: list[list[str]] = []
    for canonical, aliases_text in re.findall(r'"([^"]+)"\s*:\s*\[([^\]]*)\]', match.group(1)):
        groups.append([canonical, *re.findall(r'"([^"]+)"', aliases_text)])
    return groups


def parse_official_new_launch_items(html: str) -> list[dict[str, Any]]:
    match = OFFICIAL_NEW_LAUNCH_RE.search(html)
    return json.loads(match.group(1), strict=False) if match else []


def parse_project_basic_info(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    match = PROJECT_BASIC_INFO_RE.search(path.read_text(encoding="utf-8"))
    if not match:
        return []
    return json.loads(match.group(1), strict=False).get("rows", [])


def build_official_project(item: dict[str, Any], months: list[str]) -> dict[str, Any]:
    permits = [clean(value) for value in item.get("permits", []) if clean(value)]
    issue_dates = sorted({clean(value) for value in item.get("issueDates", []) if clean(value)})
    first_permit = permits[0] if permits else ""
    permit_id = re.search(r"\d{4}\)?\d+", first_permit)
    official_name = clean(item.get("officialProjectName"))
    detail_urls = [clean(value) for value in item.get("detailUrls", []) if clean(value)]
    residential_total = int(number(item.get("residentialTotal")) or number(item.get("approvedTotalSuites")))
    return {
        "id": f"zjw-launch-{permit_id.group(0) if permit_id else normalize_name(official_name)}",
        "group": clean(item.get("group")),
        "plate": clean(item.get("plate")),
        "project": official_name,
        "landDate": clean(item.get("landDate")),
        "status": "新取证",
        "x": 50,
        "y": 50,
        "lat": number(item.get("lat")),
        "lng": number(item.get("lng")),
        "monthly": {month: {"suites": 0, "area": 0, "price": 0, "amount": 0} for month in months},
        "suites34": 0,
        "area34": 0,
        "amount34": 0,
        "price4": 0,
        "amountAll": 0,
        "suitesAll": 0,
        "district": clean(item.get("district")),
        "address": clean(item.get("address")),
        "coordSource": "住建委坐落位置 + 业务板块近似定位",
        "coordSourceUrl": detail_urls[0] if detail_urls else "",
        "coordConfidence": clean(item.get("coordConfidence")) or "低",
        "coordSystem": "GCJ-02近似板块中心",
        "matchedName": official_name,
        "summaryRecordName": official_name,
        "summaryPresalePermit": " / ".join(permits),
        "summaryDeveloper": clean(item.get("developer")) or "住建委项目详情页",
        "officialProjectName": official_name,
        "officialResidentialTotal": residential_total,
        "officialUnsignedSuites": None,
        "officialAvailableSuites": None,
        "officialUnsignedBlueSuites": 0,
        "officialBookedSuites": None,
        "officialContractSignedSuites": None,
        "officialFilingSuites": None,
        "officialSignedStatsSuites": None,
        "officialSignedStatsArea": None,
        "officialSignedStatsAvgPrice": None,
        "officialSignedSuites": None,
        "officialDetailSignedSuites": None,
        "officialInventoryEvidenceUrl": "\n".join(detail_urls),
        "officialInventoryFetchedAt": "",
        "officialInventoryMatchStatus": "住建委新发预售证项目；克而瑞明细已匹配",
        "officialInventoryTotalAuditNote": clean(item.get("inventoryNote")),
        "approvedTotalSuites": int(number(item.get("approvedTotalSuites"))) or None,
        "source": "zjwNewLaunch",
        "presaleIssueRecords": [
            {
                "name": official_name,
                "permit": permit,
                "date": clean(item.get("issueDates", [""])[index] if index < len(item.get("issueDates", [])) else (issue_dates[0] if issue_dates else "")),
            }
            for index, permit in enumerate(permits)
        ],
        "presaleIssueDates": issue_dates,
    }


def promote_official_unmatched_projects(
    data: dict[str, Any],
    unmatched: list[dict[str, Any]],
    official_items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    existing_names = {normalize_name(project.get("project")) for project in all_projects(data)}
    by_name = {
        normalize_name(item.get("officialProjectName")): item
        for item in official_items
        if normalize_name(item.get("officialProjectName"))
    }
    promoted: list[dict[str, Any]] = []
    launch_projects = data.setdefault("launchProjects", [])
    for record in unmatched:
        probe_names = [record.get("mappedProject"), *(record.get("sourceProjects") or [])]
        item = next((by_name.get(normalize_name(name)) for name in probe_names if by_name.get(normalize_name(name))), None)
        if not item:
            continue
        project = build_official_project(item, data.get("months", []))
        key = normalize_name(project.get("project"))
        if key in existing_names:
            continue
        launch_projects.append(project)
        existing_names.add(key)
        promoted.append({
            "project": project["project"],
            "mappedProject": record.get("mappedProject"),
            "rows": record.get("rows"),
        })
    return promoted


def expanded_names(names: list[str], alias_groups: list[list[str]]) -> list[str]:
    output = [name for name in names if clean(name)]
    normalized = {normalize_name(name) for name in output if normalize_name(name)}
    for group in alias_groups:
        if any(normalize_name(name) in normalized for name in group):
            output.extend(group)
    return list(dict.fromkeys(clean(name) for name in output if clean(name)))


def project_names(project: dict[str, Any], alias_groups: list[list[str]]) -> list[str]:
    names: list[str] = []
    for field in PROJECT_FIELDS:
        names.extend(split_names(project.get(field)))
    return expanded_names(names, alias_groups)


def all_projects(data: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for collection in ("projects", "launchProjects"):
        for project in data.get(collection, []):
            marker = str(project.get("id"))
            if marker not in seen:
                seen.add(marker)
                rows.append(project)
    return rows


def read_workbook(path: Path) -> tuple[dict[str, dict[str, list[dict[str, Any]]]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [clean(value) for value in header_values]
        required = {
            "md5_str", "trade_day", "project_name", "pre_permit", "building_name",
            "unit_number", "room_number", "property_type", "layout_type", "trade_area",
            "trade_price", "trade_amount", "去化表项目名",
        }
        missing = sorted(required - set(headers))
        if missing:
            raise ValueError(f"明细文件缺少字段: {', '.join(missing)}")

        grouped: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
        raw_rows = 0
        used_rows = 0
        skipped_before_start = 0
        excluded_rows = 0
        duplicate_md5_rows = 0
        seen_md5: set[str] = set()
        source_projects: dict[str, set[str]] = defaultdict(set)
        historical_meta: dict[str, dict[str, Any]] = defaultdict(
            lambda: {"recordRows": 0, "firstDate": "", "lastDate": "", "mappedName": ""}
        )
        historical_houses: dict[str, dict[tuple[str, ...], dict[str, Any]]] = defaultdict(dict)

        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            raw_rows += 1
            property_type = clean(row.get("property_type"))
            if property_type not in ALLOWED_PROPERTY_TYPES:
                excluded_rows += 1
                continue
            month = month_label(row.get("trade_day"))
            if not month:
                excluded_rows += 1
                continue
            md5 = clean(row.get("md5_str"))
            if md5 and md5 in seen_md5:
                duplicate_md5_rows += 1
                continue
            if md5:
                seen_md5.add(md5)
            mapped_name = clean(row.get("去化表项目名")) or clean(row.get("project_name"))
            source_name = clean(row.get("project_name"))
            if not mapped_name or not source_name:
                excluded_rows += 1
                continue
            mapped_key = normalize_name(mapped_name)
            source_projects[mapped_key].add(source_name)
            trade_date = date_text(row.get("trade_day"))
            history = historical_meta[mapped_key]
            history["recordRows"] += 1
            history["mappedName"] = mapped_name
            if trade_date and (not history["firstDate"] or trade_date < history["firstDate"]):
                history["firstDate"] = trade_date
            if trade_date and trade_date > history["lastDate"]:
                history["lastDate"] = trade_date
            house_key = historical_house_key(row)
            existing_house = historical_houses[mapped_key].get(house_key)
            if not existing_house or trade_date >= existing_house["date"]:
                historical_houses[mapped_key][house_key] = {
                    "date": trade_date,
                    "area": number(row.get("trade_area")),
                    "amountWan": number(row.get("trade_amount")) / 10000,
                }
            if not is_replacement_month(month):
                skipped_before_start += 1
                continue
            grouped[month][mapped_key].append({
                "date": trade_date,
                "permit": clean(row.get("pre_permit")),
                "sourceProject": source_name,
                "building": clean(row.get("building_name")),
                "unit": clean(row.get("unit_number")),
                "room": clean(row.get("room_number")),
                "propertyType": property_type,
                "layout": clean(row.get("layout_type")),
                "area": rounded(number(row.get("trade_area")), 2),
                "unitPrice": rounded(number(row.get("trade_price")), 0),
                "totalWan": trade_amount_wan(row.get("trade_amount")),
                "_mappedName": mapped_name,
            })
            used_rows += 1

        for month_groups in grouped.values():
            for rows in month_groups.values():
                rows.sort(key=lambda item: item["date"], reverse=True)

        normalized_history = {}
        for key, meta in historical_meta.items():
            houses = list(historical_houses[key].values())
            normalized_history[key] = {
                "suites": len(houses),
                "recordRows": int(meta["recordRows"]),
                "duplicateHouseRows": int(meta["recordRows"]) - len(houses),
                "area": rounded(sum(value["area"] for value in houses), 2),
                "amountWan": rounded(sum(value["amountWan"] for value in houses), 4),
                "firstDate": meta["firstDate"],
                "lastDate": meta["lastDate"],
                "mappedName": meta["mappedName"],
            }
        history_dates = [
            value[date_key]
            for value in normalized_history.values()
            for date_key in ("firstDate", "lastDate")
            if value[date_key]
        ]
        stats = {
            "sheet": sheet.title,
            "rawRows": raw_rows,
            "usedRows": used_rows,
            "historicalUsedRows": sum(value["recordRows"] for value in normalized_history.values()),
            "historicalUniqueHouseRows": sum(value["suites"] for value in normalized_history.values()),
            "historicalDuplicateHouseRows": sum(value["duplicateHouseRows"] for value in normalized_history.values()),
            "skippedBeforeJuly2025": skipped_before_start,
            "excludedRows": excluded_rows,
            "duplicateMd5Rows": duplicate_md5_rows,
            "months": sorted(grouped, key=month_tuple),
            "mappedProjects": len(source_projects),
            "sourceProjects": {key: sorted(values) for key, values in source_projects.items()},
            "historicalMetrics": normalized_history,
            "historicalDateRange": [min(history_dates), max(history_dates)] if history_dates else [],
        }
        return grouped, stats
    finally:
        workbook.close()


def build_project_index(
    projects: list[dict[str, Any]], alias_groups: list[list[str]]
) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for project in projects:
        for name in project_names(project, alias_groups):
            key = normalize_name(name)
            if key and project not in index[key]:
                index[key].append(project)
    return index


def match_groups(
    grouped: dict[str, dict[str, list[dict[str, Any]]]],
    stats: dict[str, Any],
    projects: list[dict[str, Any]],
    alias_groups: list[list[str]],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    index = build_project_index(projects, alias_groups)
    mapped_names: dict[str, str] = {}
    for month_groups in grouped.values():
        for key, rows in month_groups.items():
            if rows:
                mapped_names.setdefault(key, clean(rows[0].get("_mappedName")))

    matches: dict[str, dict[str, Any]] = {}
    unmatched: list[dict[str, Any]] = []
    ambiguous: list[dict[str, Any]] = []
    for mapped_key, mapped_name in sorted(mapped_names.items(), key=lambda item: item[1]):
        candidate_names = [mapped_name, *stats["sourceProjects"].get(mapped_key, [])]
        candidate_names = expanded_names(candidate_names, alias_groups)
        candidate_projects: list[dict[str, Any]] = []
        for name in candidate_names:
            for part in split_names(name) or [name]:
                for project in index.get(normalize_name(part), []):
                    if project not in candidate_projects:
                        candidate_projects.append(project)
        row_count = sum(len(month_groups.get(mapped_key, [])) for month_groups in grouped.values())
        record = {
            "mappedProject": mapped_name,
            "sourceProjects": stats["sourceProjects"].get(mapped_key, []),
            "rows": row_count,
        }
        if len(candidate_projects) == 1:
            matches[mapped_key] = candidate_projects[0]
        elif not candidate_projects:
            unmatched.append(record)
        else:
            record["dashboardProjects"] = [project.get("project") for project in candidate_projects]
            ambiguous.append(record)
    return matches, unmatched, ambiguous


def attach_historical_sales(
    data: dict[str, Any],
    matches: dict[str, dict[str, Any]],
    stats: dict[str, Any],
    source_file: str,
) -> None:
    field_names = (
        "historicalTransactionSoldSuites",
        "historicalTransactionRecordRows",
        "historicalTransactionDuplicateRows",
        "historicalTransactionArea",
        "historicalTransactionAmountWan",
        "historicalTransactionStartDate",
        "historicalTransactionEndDate",
        "historicalTransactionCoverageStartDate",
        "historicalTransactionCoverageEndDate",
        "historicalTransactionSource",
        "historicalTransactionScope",
        "historicalTransactionMappedNames",
    )
    for project in all_projects(data):
        for field in field_names:
            project.pop(field, None)

    aggregated: dict[str, dict[str, Any]] = {}
    for mapped_key, project in matches.items():
        history = stats.get("historicalMetrics", {}).get(mapped_key)
        if not history:
            continue
        marker = str(project.get("id"))
        target = aggregated.setdefault(marker, {
            "project": project,
            "suites": 0,
            "recordRows": 0,
            "duplicateRows": 0,
            "area": 0.0,
            "amountWan": 0.0,
            "firstDate": "",
            "lastDate": "",
            "mappedNames": [],
        })
        target["suites"] += int(history.get("suites") or 0)
        target["recordRows"] += int(history.get("recordRows") or 0)
        target["duplicateRows"] += int(history.get("duplicateHouseRows") or 0)
        target["area"] += number(history.get("area"))
        target["amountWan"] += number(history.get("amountWan"))
        first_date = clean(history.get("firstDate"))
        last_date = clean(history.get("lastDate"))
        if first_date and (not target["firstDate"] or first_date < target["firstDate"]):
            target["firstDate"] = first_date
        if last_date and last_date > target["lastDate"]:
            target["lastDate"] = last_date
        mapped_name = clean(history.get("mappedName"))
        if mapped_name and mapped_name not in target["mappedNames"]:
            target["mappedNames"].append(mapped_name)

    for target in aggregated.values():
        project = target["project"]
        project["historicalTransactionSoldSuites"] = int(target["suites"])
        project["historicalTransactionRecordRows"] = int(target["recordRows"])
        project["historicalTransactionDuplicateRows"] = int(target["duplicateRows"])
        project["historicalTransactionArea"] = rounded(target["area"], 2)
        project["historicalTransactionAmountWan"] = rounded(target["amountWan"], 4)
        project["historicalTransactionStartDate"] = target["firstDate"]
        project["historicalTransactionEndDate"] = target["lastDate"]
        coverage = stats.get("historicalDateRange", [])
        project["historicalTransactionCoverageStartDate"] = coverage[0] if coverage else ""
        project["historicalTransactionCoverageEndDate"] = coverage[-1] if coverage else ""
        project["historicalTransactionSource"] = source_file
        project["historicalTransactionScope"] = "仅统计2025年1月至今普通住宅/别墅；同一房源多次成交按最新一笔去重计数"
        project["historicalTransactionMappedNames"] = target["mappedNames"]


def audit_historical_sales(data: dict[str, Any]) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for project in all_projects(data):
        sold = project.get("historicalTransactionSoldSuites")
        sold = int(number(sold)) if sold is not None else None
        record_rows = int(number(project.get("historicalTransactionRecordRows")))
        duplicate_rows = int(number(project.get("historicalTransactionDuplicateRows")))
        rows.append({
            "project": project.get("project"),
            "historicalSoldSuites": sold,
            "historicalRecordRows": record_rows,
            "duplicateHouseRows": duplicate_rows,
            "status": "已匹配" if sold is not None else "历史明细未匹配",
            "historyStartDate": project.get("historicalTransactionStartDate", ""),
            "historyEndDate": project.get("historicalTransactionEndDate", ""),
        })
    return {
        "projects": len(rows),
        "matchedProjects": sum(row["status"] == "已匹配" for row in rows),
        "uniqueSoldSuites": sum(row["historicalSoldSuites"] or 0 for row in rows),
        "recordRows": sum(row["historicalRecordRows"] for row in rows),
        "duplicateHouseRows": sum(row["duplicateHouseRows"] for row in rows),
        "issues": [row for row in rows if row["status"] != "已匹配"],
        "rows": rows,
    }


def build_transaction_payload(
    grouped: dict[str, dict[str, list[dict[str, Any]]]],
    matches: dict[str, dict[str, Any]],
    alias_groups: list[list[str]],
    source_file: str,
) -> dict[str, Any]:
    months: dict[str, Any] = {}
    for month in sorted(grouped, key=month_tuple):
        projects: dict[str, Any] = {}
        aliases: dict[str, str] = {}
        for mapped_key, rows in grouped[month].items():
            project = matches.get(mapped_key)
            if not project:
                continue
            primary = normalize_name(project.get("project"))
            clean_rows = []
            for row in rows:
                item = dict(row)
                item.pop("_mappedName", None)
                clean_rows.append(item)
            projects[primary] = {
                "projectName": project.get("project", ""),
                "rawProjectName": project.get("project", ""),
                "cricProjectName": clean_rows[0].get("sourceProject", "") if clean_rows else "",
                "matchedProjectName": rows[0].get("_mappedName", "") if rows else "",
                "plate": project.get("plate", ""),
                "group": project.get("group", ""),
                "rows": clean_rows,
                "summary": summarize_rows(clean_rows),
            }
            alias_names = [
                *project_names(project, alias_groups),
                rows[0].get("_mappedName", "") if rows else "",
                *(row.get("sourceProject", "") for row in clean_rows),
            ]
            for name in alias_names:
                alias_key = normalize_name(name)
                if alias_key and alias_key != primary:
                    aliases[alias_key] = primary
        months[month] = {
            "projects": projects,
            "aliases": aliases,
            "summary": {
                "projects": len(projects),
                "rows": sum(len(project["rows"]) for project in projects.values()),
            },
        }
    return {
        "source": f"{source_file}（新明细全量替换）",
        "sheet": "Sheet1",
        "scope": "普通住宅/别墅，按克而瑞新明细逐套聚合",
        "months": months,
        "summary": {
            "months": len(months),
            "projects": sum(month["summary"]["projects"] for month in months.values()),
            "rows": sum(month["summary"]["rows"] for month in months.values()),
            "skippedRows": 0,
            "excludedParkingRows": 0,
        },
    }


def write_transaction_js(path: Path, payload: dict[str, Any]) -> None:
    text = (
        "window.TRANSACTION_DETAILS = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
        + "window.MAY_TRANSACTION_DETAILS = window.TRANSACTION_DETAILS.months['26年5月'] || "
        + "{projects:{},aliases:{},summary:{projects:0,rows:0}};\n"
    )
    path.write_text(text, encoding="utf-8")


def recalc_project(project: dict[str, Any], months: list[str]) -> None:
    monthly = project.setdefault("monthly", {})
    recent = months[-2:]
    project["suites34"] = rounded(sum(number(monthly.get(month, {}).get("suites")) for month in recent))
    project["area34"] = rounded(sum(number(monthly.get(month, {}).get("area")) for month in recent), 2)
    project["amount34"] = rounded(sum(number(monthly.get(month, {}).get("amount")) for month in recent), 4)
    project["price4"] = monthly.get(months[-1], {}).get("price", 0) if months else 0
    project["suitesAll"] = rounded(sum(number(metric.get("suites")) for metric in monthly.values()))
    project["amountAll"] = rounded(sum(number(metric.get("amount")) for metric in monthly.values()), 4)


def aggregate_month(projects: list[dict[str, Any]], month: str) -> dict[str, Any]:
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects)
    return {
        "month": month,
        "suites": rounded(suites),
        "area": rounded(area, 2),
        "amount": rounded(amount, 4),
        "price": int(round(amount * 10000 / area)) if area else 0,
    }


def refresh_data(
    data: dict[str, Any],
    payload: dict[str, Any],
    matches: dict[str, dict[str, Any]],
) -> None:
    replacement_months = sorted(payload["months"], key=month_tuple)
    data["months"] = [month for month in data.get("months", []) if not is_replacement_month(month)] + replacement_months
    data["months"] = list(dict.fromkeys(data["months"]))
    project_rows = all_projects(data)
    by_project_month: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for month, month_data in payload["months"].items():
        for key, group in month_data.get("projects", {}).items():
            by_project_month[key][month] = group["summary"]

    for project in project_rows:
        primary = normalize_name(project.get("project"))
        monthly = project.setdefault("monthly", {})
        for month in replacement_months:
            summary = by_project_month.get(primary, {}).get(month)
            monthly[month] = {
                "suites": summary["suites"] if summary else 0,
                "area": summary["area"] if summary else 0,
                "price": summary["avgPrice"] if summary else 0,
                "amount": summary["amountWan"] if summary else 0,
            }
        recalc_project(project, data["months"])

    dashboard_projects = data.get("projects", [])
    recent = data["months"][-2:]
    suites = sum(number(project["monthly"][month]["suites"]) for project in dashboard_projects for month in recent)
    area = sum(number(project["monthly"][month]["area"]) for project in dashboard_projects for month in recent)
    amount = sum(number(project["monthly"][month]["amount"]) for project in dashboard_projects for month in recent)
    data["totals"] = {
        "projects": len(dashboard_projects),
        "plates": len({project.get("plate") for project in dashboard_projects}),
        "active": sum(project.get("status") == "在售" for project in dashboard_projects),
        "suites34": rounded(suites),
        "amount34": rounded(amount, 4),
        "area34": rounded(area, 2),
        "avgPrice34": int(round(amount * 10000 / area)) if area else 0,
    }
    data["monthlyTotals"] = [aggregate_month(dashboard_projects, month) for month in data["months"]]
    group_totals = []
    for group in dict.fromkeys(project.get("group") for project in dashboard_projects):
        group_projects = [project for project in dashboard_projects if project.get("group") == group]
        group_totals.append({
            "group": group,
            "projects": len(group_projects),
            "suites34": rounded(sum(number(project.get("suites34")) for project in group_projects)),
            "amount34": rounded(sum(number(project.get("amount34")) for project in group_projects), 4),
            "color": data.get("colors", {}).get(group, "#8ea0b8"),
        })
    data["groupTotals"] = group_totals
    data["colors"] = {row["group"]: row["color"] for row in group_totals}
    policy = data.setdefault("sourcePolicy", {})
    policy["成交明细口径"] = "普通住宅/别墅；2025年7月起以项目成交明细.xlsx为唯一事实源"
    policy["成交校验"] = "月度套数、面积、金额和均价均由逐套明细重算并闭合"
    policy.pop("剩余套数口径", None)
    policy["历史已售口径"] = "仅统计2025年1月至今；普通住宅/别墅；同一房源多次成交按最新一笔去重计数"
    for month in replacement_months:
        full = month_tuple(month)
        policy[f"{full[0]}年{full[1]}月"] = "克而瑞项目成交明细逐套聚合（新明细全量替换）"
        policy[f"{month}成交明细"] = "克而瑞项目成交明细（普通住宅/别墅）"


def update_html(html: str, data: dict[str, Any], data_suffix: str, max_date: str) -> str:
    replacement = "const DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n" + data_suffix
    html = DATA_RE.sub(lambda _: replacement, html, count=1)
    html = re.sub(
        r'<script src="transaction_details\.js[^"]*"></script>',
        '<script src="transaction_details.js?v=20260814-authoritative"></script>',
        html,
        count=1,
    )
    html = re.sub(
        r'\n\s*<script src="(?:june|july|new_launch)_transaction_details\.js[^"]*"></script>',
        "",
        html,
    )
    html = re.sub(
        r'(<div class="cutoff"><i>▣</i><span>数据截至：</span><b>).*?(</b></div>)',
        rf"\g<1>{max_date.replace('/', '-')}\g<2>",
        html,
        count=1,
    )
    html = re.sub(
        r"const periodOptions = .*?;",
        "const periodOptions = "
        + json.dumps([[month, month] for month in reversed(data["months"])], ensure_ascii=False)
        + ";",
        html,
        count=1,
    )
    html = re.sub(
        r"const recentLaunchMonths = \[[^\]]*\]",
        "const recentLaunchMonths = " + json.dumps(data["months"][-2:], ensure_ascii=False),
        html,
        count=1,
    )
    return html


def closure_report(data: dict[str, Any], payload: dict[str, Any]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    payload_lookup = {
        (month, key): group["summary"]
        for month, month_data in payload["months"].items()
        for key, group in month_data.get("projects", {}).items()
    }
    for project in all_projects(data):
        key = normalize_name(project.get("project"))
        for month in payload["months"]:
            detail = payload_lookup.get((month, key), {"suites": 0, "area": 0, "amountWan": 0, "avgPrice": 0})
            metric = project.get("monthly", {}).get(month, {})
            if (
                number(metric.get("suites")) != number(detail.get("suites"))
                or abs(number(metric.get("area")) - number(detail.get("area"))) > 0.01
                or abs(number(metric.get("amount")) - number(detail.get("amountWan"))) > 0.01
                or number(metric.get("price")) != number(detail.get("avgPrice"))
            ):
                issues.append({"project": project.get("project"), "month": month, "dashboard": metric, "detail": detail})
    return issues


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--html", type=Path, default=Path("index.html"))
    parser.add_argument("--detail-js", type=Path, default=Path("transaction_details.js"))
    parser.add_argument("--report", type=Path, default=Path("transaction_history_replace_report.json"))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--allow-unmatched", action="store_true")
    args = parser.parse_args()

    html = args.html.read_text(encoding="utf-8")
    data_match = DATA_RE.search(html)
    if not data_match:
        raise RuntimeError("DATA block not found")
    data = json.loads(data_match.group(1), strict=False)
    alias_groups = parse_alias_groups(html)
    grouped, stats = read_workbook(args.xlsx)
    projects = all_projects(data)
    matches, unmatched, ambiguous = match_groups(grouped, stats, projects, alias_groups)
    promoted_projects = promote_official_unmatched_projects(
        data,
        unmatched,
        parse_official_new_launch_items(html),
    )
    if promoted_projects:
        projects = all_projects(data)
        matches, unmatched, ambiguous = match_groups(grouped, stats, projects, alias_groups)
    attach_historical_sales(data, matches, stats, args.xlsx.name)
    payload = build_transaction_payload(grouped, matches, alias_groups, args.xlsx.name)
    refresh_data(data, payload, matches)
    closure_issues = closure_report(data, payload)
    historical_sales_audit = audit_historical_sales(data)
    max_date = max(
        row["date"]
        for month_data in payload["months"].values()
        for project in month_data["projects"].values()
        for row in project["rows"]
    )
    report = {
        "source": str(args.xlsx),
        "stats": stats,
        "matchedMappedProjects": len(matches),
        "matchedRows": payload["summary"]["rows"],
        "matchedHistoricalUniqueHouses": sum(
            stats.get("historicalMetrics", {}).get(key, {}).get("suites", 0)
            for key in matches
        ),
        "matchedHistoricalRecordRows": sum(
            stats.get("historicalMetrics", {}).get(key, {}).get("recordRows", 0)
            for key in matches
        ),
        "matchedHistoricalDuplicateHouseRows": sum(
            stats.get("historicalMetrics", {}).get(key, {}).get("duplicateHouseRows", 0)
            for key in matches
        ),
        "promotedOfficialProjects": promoted_projects,
        "unmatched": unmatched,
        "ambiguous": ambiguous,
        "replacementMonths": list(payload["months"]),
        "monthTotals": [
            {
                "month": month,
                "rows": month_data["summary"]["rows"],
                "projects": month_data["summary"]["projects"],
            }
            for month, month_data in payload["months"].items()
        ],
        "closureIssues": closure_issues,
        "historicalSalesAudit": historical_sales_audit,
        "maxTradeDate": max_date,
    }
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    if (unmatched or ambiguous) and not args.allow_unmatched and not args.dry_run:
        raise RuntimeError("存在未匹配或多义项目，已停止写入；先查看核对报告")
    if closure_issues:
        raise RuntimeError("明细与月度汇总未闭合，已停止写入")
    if not args.dry_run:
        write_transaction_js(args.detail_js, payload)
        updated_html = update_html(html, data, data_match.group(2), max_date)
        args.html.write_text(updated_html, encoding="utf-8")

    print(json.dumps({
        "dryRun": args.dry_run,
        "sourceRows": stats["rawRows"],
        "replacementRows": payload["summary"]["rows"],
        "matchedMappedProjects": len(matches),
        "promotedOfficialProjects": len(promoted_projects),
        "unmatchedProjects": len(unmatched),
        "ambiguousProjects": len(ambiguous),
        "closureIssues": len(closure_issues),
        "replacementMonths": list(payload["months"]),
        "maxTradeDate": max_date,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

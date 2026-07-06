#!/usr/bin/env python3
"""Update the Beijing dashboard with June CRIC summary/detail data.

Inputs:
- One CRIC June project cumulative summary workbook.
- A folder of project-level June transaction detail workbooks.

The dashboard monthly metric uses the CRIC summary workbook for matched projects.
Detail rows are written to a standalone June transaction-detail JS file so the
project detail drawer can show June rows.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATA_RE = re.compile(
    r"const DATA = (.*?);\n(const LAUNCH_OFFICIAL_INVENTORY_OVERRIDES|const DEFAULT_PERIODS)",
    re.S,
)
MONTH = "26年6月"
MONTH_FULL = "2026年6月"
DETAIL_GLOBAL_RE = re.compile(r'(<script src="transaction_details\.js[^"]*"></script>)')


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).lower()
    replacements = {
        "雲": "云",
        "澐": "云",
        "樹": "树",
        "華": "华",
        "鳴": "鸣",
        "號": "号",
        "鄕": "乡",
        "臺": "台",
        "橒": "云",
        "叁": "三",
        "贰": "二",
        "壹": "一",
        "玖": "九",
        "·": "",
        "•": "",
        ".": "",
        "。": "",
        "（": "(",
        "）": ")",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[\s\-—_/\\()（）【】\[\]《》“”\"'：:；;，,]+", "", text)


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def rounded(value: float, digits: int = 2) -> int | float:
    value = round(float(value), digits)
    if abs(value - round(value)) < 0.000001:
        return int(round(value))
    return value


def row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}


def find_header_row(rows: list[tuple[Any, ...]], required: str = "项目名称") -> int:
    for index, row in enumerate(rows):
        if required in [clean(value) for value in row]:
            return index
    raise ValueError(f"Cannot find header row with {required}")


def load_summary(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header_index = find_header_row(rows, "项目名称")
        headers = [clean(value) for value in rows[header_index]]
        records: dict[str, dict[str, Any]] = {}
        for values in rows[header_index + 1:]:
            row = row_dict(headers, values)
            project = clean(row.get("项目名称"))
            if not project:
                continue
            area = number(row.get("成交面积(㎡)"))
            suites = int(round(number(row.get("成交套数(套)"))))
            amount_wan = number(row.get("成交金额(元)")) / 10000
            price = int(round(number(row.get("成交均价(元/㎡)")))) if number(row.get("成交均价(元/㎡)")) else 0
            records[normalize_name(project)] = {
                "sourceProject": project,
                "district": clean(row.get("区域")),
                "cricPlate": clean(row.get("板块")),
                "developer": clean(row.get("开发商")),
                "suites": suites,
                "area": rounded(area, 2),
                "price": price,
                "amount": rounded(amount_wan, 4),
            }
        return records
    finally:
        wb.close()


def detail_file_paths(folder: Path) -> list[Path]:
    return [
        path for path in sorted(folder.glob("*.xlsx"))
        if not path.name.startswith(("~", ".~")) and path.is_file()
    ]


def read_detail_file(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return []
        header_index = find_header_row(rows, "成交时间")
        headers = [clean(value) for value in rows[header_index]]
        output: list[dict[str, Any]] = []
        for values in rows[header_index + 1:]:
            row = row_dict(headers, values)
            project = clean(row.get("项目名称"))
            date = clean(row.get("成交时间"))
            if not project or not date:
                continue
            property_type = clean(row.get("物业类型"))
            if property_type and property_type not in {"普通住宅", "别墅"}:
                continue
            total_wan = number(row.get("成交总价(万元)")) or number(row.get("成交总价(元)")) / 10000
            output.append(
                {
                    "date": date,
                    "permit": clean(row.get("预售证号")),
                    "sourceProject": project,
                    "building": clean(row.get("楼栋名称")),
                    "unit": clean(row.get("单元号")),
                    "room": clean(row.get("室号")),
                    "propertyType": property_type,
                    "layout": clean(row.get("房型")),
                    "area": rounded(number(row.get("成交面积(㎡)")), 2),
                    "unitPrice": rounded(number(row.get("成交单价(元/㎡)")), 0),
                    "totalWan": rounded(total_wan, 2),
                    "_sourceFile": path.name,
                }
            )
        return output
    finally:
        wb.close()


def load_details(folder: Path) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for path in detail_file_paths(folder):
        for row in read_detail_file(path):
            key = normalize_name(row["sourceProject"])
            group = grouped.setdefault(
                key,
                {
                    "sourceProject": row["sourceProject"],
                    "rows": [],
                    "sourceFiles": set(),
                },
            )
            group["rows"].append(row)
            group["sourceFiles"].add(row["_sourceFile"])

    for group in grouped.values():
        rows = group["rows"]
        rows.sort(key=lambda item: item["date"], reverse=True)
        area = sum(number(row["area"]) for row in rows)
        amount = sum(number(row["totalWan"]) for row in rows)
        group["summary"] = {
            "suites": len(rows),
            "area": rounded(area, 2),
            "amountWan": rounded(amount, 2),
            "avgPrice": int(round(amount * 10000 / area)) if area else 0,
        }
        group["sourceFiles"] = sorted(group["sourceFiles"])
        for row in rows:
            row.pop("_sourceFile", None)
    return grouped


def project_candidate_names(project: dict[str, Any]) -> list[str]:
    fields = [
        "cricProjectName",
        "janAprMatchedName",
        "junMatchedName",
        "junCricProjectName",
        "matchedName",
        "project",
        "summaryRecordName",
        "officialProjectName",
    ]
    manual_aliases = {
        "北投栖澐湾": ["北投·云帆汀澜", "北投云帆汀澜"],
        "中海玖樹满和": ["中海·九树满和", "中海九树满和"],
        "中建方程国贤府": ["方程国贤府", "中建·方程国贤府"],
        "未来城星寰时代": ["未来城·星寰时代"],
    }
    names: list[str] = []
    for field in fields:
        value = clean(project.get(field))
        if not value:
            continue
        names.extend(re.split(r"[；;、/\n]+", value))
    names.extend(manual_aliases.get(clean(project.get("project")), []))
    output: list[str] = []
    for name in names:
        name = clean(name)
        if name and name not in output:
            output.append(name)
    return output


def match_record(records: dict[str, Any], project: dict[str, Any]) -> tuple[str, Any | None, str]:
    for name in project_candidate_names(project):
        key = normalize_name(name)
        if key and key in records:
            return key, records[key], name
    return "", None, ""


def all_dashboard_projects(data: dict[str, Any]) -> list[dict[str, Any]]:
    projects: list[dict[str, Any]] = []
    seen: set[int] = set()
    for collection_name in ("projects", "launchProjects"):
        for project in data.get(collection_name, []):
            marker = id(project)
            if marker not in seen:
                seen.add(marker)
                projects.append(project)
    return projects


def calc_price(amount: float, area: float) -> int:
    return int(round(amount * 10000 / area)) if area else 0


def infer_custom_plate_and_group(data: dict[str, Any], summary_row: dict[str, Any]) -> tuple[str, str]:
    raw_plate = clean(summary_row.get("cricPlate")).replace("板块", "")
    raw_plate = raw_plate or clean(summary_row.get("district")) or "待归类"
    existing_by_key: dict[str, tuple[str, str]] = {}
    for project in data.get("projects", []):
        plate = clean(project.get("plate"))
        if plate:
            existing_by_key[normalize_name(plate)] = (plate, clean(project.get("group")) or "未分组")
    if normalize_name(raw_plate) in existing_by_key:
        return existing_by_key[normalize_name(raw_plate)]
    return raw_plate, "未分组"


def inferred_location(data: dict[str, Any], plate: str) -> dict[str, Any]:
    same_plate = [
        project for project in data.get("projects", [])
        if clean(project.get("plate")) == plate
    ]
    with_xy = [project for project in same_plate if number(project.get("x")) or number(project.get("y"))]
    with_coord = [
        project for project in same_plate
        if isinstance(project.get("lat"), (int, float)) and isinstance(project.get("lng"), (int, float))
    ]
    location: dict[str, Any] = {}
    if with_xy:
        location["x"] = rounded(sum(number(project.get("x")) for project in with_xy) / len(with_xy), 2)
        location["y"] = rounded(sum(number(project.get("y")) for project in with_xy) / len(with_xy), 2)
    else:
        location["x"] = 50
        location["y"] = 50
    if with_coord:
        location["lat"] = rounded(sum(number(project.get("lat")) for project in with_coord) / len(with_coord), 6)
        location["lng"] = rounded(sum(number(project.get("lng")) for project in with_coord) / len(with_coord), 6)
        location["coordSource"] = "同板块已核验项目坐标均值（待补项目精确坐标）"
        location["coordConfidence"] = "低"
        return location
    center = data.get("plateCenters", {}).get(plate, {})
    if isinstance(center.get("lat"), (int, float)) and isinstance(center.get("lng"), (int, float)):
        location["lat"] = center["lat"]
        location["lng"] = center["lng"]
        location["coordSource"] = "板块中心临时定位（待补项目精确坐标）"
        location["coordConfidence"] = "低"
    return location


def add_new_detail_projects(
    data: dict[str, Any],
    summary: dict[str, dict[str, Any]],
    details: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    existing_keys = {
        normalize_name(name)
        for project in all_dashboard_projects(data)
        for name in project_candidate_names(project)
        if normalize_name(name)
    }
    numeric_ids = [
        int(project["id"])
        for project in all_dashboard_projects(data)
        if isinstance(project.get("id"), int) or str(project.get("id", "")).isdigit()
    ]
    max_id = max(numeric_ids or [0])
    added: list[dict[str, Any]] = []
    for key, detail in details.items():
        if key in existing_keys or not detail.get("summary", {}).get("suites"):
            continue
        summary_row = summary.get(key)
        if not summary_row or not summary_row.get("suites"):
            continue
        max_id += 1
        plate, group = infer_custom_plate_and_group(data, summary_row)
        project_name = summary_row["sourceProject"]
        monthly = {month: {"suites": 0, "area": 0, "price": 0, "amount": 0} for month in data["months"]}
        monthly[MONTH] = {
            "suites": summary_row["suites"],
            "area": summary_row["area"],
            "price": summary_row["price"],
            "amount": summary_row["amount"],
        }
        project = {
            "id": max_id,
            "group": group,
            "plate": plate,
            "project": project_name,
            "landDate": "",
            "status": "在售",
            "monthly": monthly,
            "district": summary_row.get("district", ""),
            "address": "",
            "matchedName": project_name,
            "cricProjectName": project_name,
            "summaryDeveloper": summary_row.get("developer", ""),
            "junDataSource": "克尔瑞",
            "junMatchedName": project_name,
            "junCricProjectName": project_name,
            "isNewFromJuneCric": True,
        }
        project.update(inferred_location(data, plate))
        data.setdefault("projects", []).append(project)
        existing_keys.add(key)
        added.append({"project": project_name, "summaryProject": project_name, "detailProject": detail["sourceProject"]})
    return added


def recalc_project(project: dict[str, Any], months: list[str]) -> None:
    recent_months = months[-2:]
    monthly = project.get("monthly", {})
    project["suites34"] = rounded(sum(number(monthly.get(month, {}).get("suites")) for month in recent_months))
    project["area34"] = rounded(sum(number(monthly.get(month, {}).get("area")) for month in recent_months), 2)
    project["amount34"] = rounded(sum(number(monthly.get(month, {}).get("amount")) for month in recent_months), 4)
    latest = monthly.get(recent_months[-1], {}) if recent_months else {}
    project["price4"] = latest.get("price", 0)
    project["suitesAll"] = rounded(sum(number(metric.get("suites")) for metric in monthly.values()))
    project["amountAll"] = rounded(sum(number(metric.get("amount")) for metric in monthly.values()), 4)


def aggregate_month(projects: list[dict[str, Any]], month: str) -> dict[str, Any]:
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects)
    return {"month": month, "suites": rounded(suites), "area": rounded(area, 2), "amount": rounded(amount, 4), "price": calc_price(amount, area)}


def refresh_aggregates(data: dict[str, Any]) -> None:
    months = data["months"]
    for collection_name in ("projects", "launchProjects"):
        for project in data.get(collection_name, []):
            recalc_project(project, months)

    projects = data["projects"]
    recent = months[-2:]
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects for month in recent)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects for month in recent)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects for month in recent)
    data["totals"] = {
        "projects": len(projects),
        "plates": len({project.get("plate") for project in projects}),
        "active": sum(project.get("status") == "在售" for project in projects),
        "suites34": rounded(suites),
        "amount34": rounded(amount, 4),
        "area34": rounded(area, 2),
        "avgPrice34": calc_price(amount, area),
    }
    data["monthlyTotals"] = [aggregate_month(projects, month) for month in months]

    group_totals = []
    for group in dict.fromkeys(project.get("group") for project in projects):
        group_projects = [project for project in projects if project.get("group") == group]
        group_totals.append(
            {
                "group": group,
                "projects": len(group_projects),
                "suites34": rounded(sum(number(project.get("suites34")) for project in group_projects)),
                "amount34": rounded(sum(number(project.get("amount34")) for project in group_projects), 4),
                "color": data.get("colors", {}).get(group, "#8ea0b8"),
            }
        )
    data["groupTotals"] = group_totals
    data["colors"] = {row["group"]: row["color"] for row in group_totals}
    policy = data.setdefault("sourcePolicy", {})
    policy["2026年6月"] = "克尔瑞项目累计供求汇总（商品住宅）"
    policy["26年6月成交明细"] = "克尔瑞项目成交明细（普通住宅/别墅）"


def ensure_june_controls(html: str, data: dict[str, Any]) -> str:
    full = "2026年6月"
    period_options = list(reversed(data["months"]))
    html = re.sub(
        r'(<div class="cutoff"><i>▣</i><span>数据截至：</span><b>).*?(</b></div>)',
        rf"\g<1>{full}\g<2>",
        html,
        count=1,
    )
    html = re.sub(
        r"const DEFAULT_PERIODS = \[[^\]]*\];",
        "const DEFAULT_PERIODS = [];",
        html,
        count=1,
    )
    html = re.sub(
        r'const latestSortMonth = .*?;',
        'const latestSortMonth = DATA.months.filter(monthParts).sort((a,b)=>monthRank(b)-monthRank(a))[0] || "26年6月";',
        html,
        count=1,
    )
    html = re.sub(
        r"const periodOptions = .*?;",
        "const periodOptions = "
        + json.dumps([[month, month] for month in period_options], ensure_ascii=False)
        + ";",
        html,
        count=1,
    )
    html = re.sub(
        r"const recentLaunchMonths = \[[^\]]*\]",
        "const recentLaunchMonths = " + json.dumps(data["months"][-2:], ensure_ascii=False),
        html,
        count=1,
    )
    if '<script src="june_transaction_details.js"></script>' not in html:
        html = DETAIL_GLOBAL_RE.sub(
            r'\1\n  <script src="june_transaction_details.js"></script>',
            html,
            count=1,
        )
    html = html.replace(
        "26年5月沿用原页面明细，已排除车库/车位",
        "26年6月为克尔瑞项目明细；1–5月沿用原页面明细，已排除车库/车位",
    )
    html = html.replace(
        '  if (month === "26年5月") {\n'
        '    const label = /克[而尔]瑞/.test(String(p.mayDataSource || "")) ? "克尔瑞" : "天朗";\n'
        '    return { label };\n'
        '  }\n'
        '  return { label:"天朗" };\n',
        '  if (month === "26年5月") {\n'
        '    const label = /克[而尔]瑞/.test(String(p.mayDataSource || "")) ? "克尔瑞" : "天朗";\n'
        '    return { label };\n'
        '  }\n'
        '  if (month === "26年6月") {\n'
        '    return { label:"克尔瑞" };\n'
        '  }\n'
        '  return { label:"天朗" };\n',
    )
    return html


def build_june_detail_payload(data: dict[str, Any], details: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    month_projects: dict[str, Any] = {}
    aliases: dict[str, str] = {}
    unmatched_detail: list[dict[str, Any]] = []

    dashboard_projects = all_dashboard_projects(data)
    for project in dashboard_projects:
        detail_key, detail, matched_name = match_record(details, project)
        if not detail:
            continue
        primary = normalize_name(project["project"])
        rows = detail["rows"]
        summary = detail["summary"]
        group = {
            "projectName": project["project"],
            "rawProjectName": project["project"],
            "cricProjectName": detail["sourceProject"],
            "matchedProjectName": matched_name or project["project"],
            "plate": project.get("plate", ""),
            "group": project.get("group", ""),
            "rows": rows,
            "summary": summary,
        }
        month_projects[primary] = group
        for alias in [project.get("project"), detail["sourceProject"], project.get("cricProjectName"), project.get("matchedName"), matched_name]:
            key = normalize_name(alias)
            if key and key != primary:
                aliases[key] = primary

    used_detail_keys = {match_record(details, p)[0] for p in dashboard_projects if match_record(details, p)[1]}
    for key, detail in details.items():
        if key not in used_detail_keys and detail["summary"]["suites"]:
            unmatched_detail.append({
                "detailProject": detail["sourceProject"],
                "suites": detail["summary"]["suites"],
                "area": detail["summary"]["area"],
                "amountWan": detail["summary"]["amountWan"],
                "sourceFiles": detail["sourceFiles"],
            })

    payload = {
        "source": "华北外部数据/6月项目明细",
        "sheet": "CRIC-北京-项目详情.交易.项目累计.成交明细",
        "scope": "普通住宅/别墅，已排除车库/车位",
        "month": MONTH_FULL,
        "projects": month_projects,
        "aliases": aliases,
        "summary": {
            "projects": len(month_projects),
            "rows": sum(len(project["rows"]) for project in month_projects.values()),
            "excludedParkingRows": 0,
        },
    }
    return payload, {"unmatchedDetail": unmatched_detail}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--html", type=Path, default=Path("index.html"))
    parser.add_argument("--summary", type=Path, required=True)
    parser.add_argument("--detail-dir", type=Path, required=True)
    parser.add_argument("--detail-js", type=Path, default=Path("june_transaction_details.js"))
    parser.add_argument("--diff-json", type=Path, default=Path("june_cric_diff.json"))
    args = parser.parse_args()

    html = args.html.read_text(encoding="utf-8")
    match = DATA_RE.search(html)
    if not match:
        raise RuntimeError("DATA block not found")
    data = json.loads(match.group(1), strict=False)
    if MONTH not in data["months"]:
        data["months"].append(MONTH)
    summary = load_summary(args.summary)
    details = load_details(args.detail_dir)
    added_projects = add_new_detail_projects(data, summary, details)

    matched_projects = 0
    unmatched_dashboard: list[dict[str, Any]] = []
    diff_rows: list[dict[str, Any]] = []
    main_project_ids = {id(project) for project in data.get("projects", [])}

    for project in all_dashboard_projects(data):
        summary_key, summary_row, summary_match = match_record(summary, project)
        detail_key, detail_row, detail_match = match_record(details, project)
        if summary_row:
            if id(project) in main_project_ids:
                matched_projects += 1
            project.setdefault("monthly", {})[MONTH] = {
                "suites": summary_row["suites"],
                "area": summary_row["area"],
                "price": summary_row["price"],
                "amount": summary_row["amount"],
            }
            project["junDataSource"] = "克尔瑞"
            project["junMatchedName"] = summary_match
            project["junCricProjectName"] = summary_row["sourceProject"]
        else:
            project.setdefault("monthly", {})[MONTH] = {"suites": 0, "area": 0, "price": 0, "amount": 0}
            if id(project) in main_project_ids:
                unmatched_dashboard.append({"project": project["project"], "candidates": project_candidate_names(project)})
        if id(project) in main_project_ids and (summary_row or detail_row):
            ds = detail_row["summary"] if detail_row else {"suites": 0, "area": 0, "amountWan": 0, "avgPrice": 0}
            ss = summary_row if summary_row else {"suites": 0, "area": 0, "amount": 0, "price": 0, "sourceProject": ""}
            suite_diff = int(round(number(ss.get("suites")) - number(ds.get("suites"))))
            area_diff = round(number(ss.get("area")) - number(ds.get("area")), 2)
            amount_diff = round(number(ss.get("amount")) - number(ds.get("amountWan")), 2)
            if suite_diff or abs(area_diff) > 0.05 or abs(amount_diff) > 0.05 or (summary_row and not detail_row) or (detail_row and not summary_row):
                diff_rows.append(
                    {
                        "dashboardProject": project["project"],
                        "summaryProject": ss.get("sourceProject", ""),
                        "detailProject": detail_row.get("sourceProject", "") if detail_row else "",
                        "summarySuites": ss.get("suites", 0),
                        "detailSuites": ds.get("suites", 0),
                        "suiteDiff": suite_diff,
                        "summaryArea": ss.get("area", 0),
                        "detailArea": ds.get("area", 0),
                        "areaDiff": area_diff,
                        "summaryAmountWan": ss.get("amount", 0),
                        "detailAmountWan": ds.get("amountWan", 0),
                        "amountDiff": amount_diff,
                        "summaryMatchedBy": summary_match,
                        "detailMatchedBy": detail_match,
                    }
                )

    refresh_aggregates(data)
    replacement = "const DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n" + match.group(2)
    html = DATA_RE.sub(lambda _: replacement, html, count=1)
    html = ensure_june_controls(html, data)
    args.html.write_text(html, encoding="utf-8")

    detail_payload, detail_extra = build_june_detail_payload(data, details)
    merge_js = """
if (window.TRANSACTION_DETAILS) {
  window.TRANSACTION_DETAILS.months = window.TRANSACTION_DETAILS.months || {};
  window.TRANSACTION_DETAILS.months["26年6月"] = window.JUNE_TRANSACTION_DETAILS;
  window.TRANSACTION_DETAILS.summary = window.TRANSACTION_DETAILS.summary || {};
  window.TRANSACTION_DETAILS.summary.months = Object.keys(window.TRANSACTION_DETAILS.months).length;
}
""".strip()
    args.detail_js.write_text(
        "window.JUNE_TRANSACTION_DETAILS = "
        + json.dumps(detail_payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
        + merge_js
        + "\n",
        encoding="utf-8",
    )

    used_summary_keys = {match_record(summary, p)[0] for p in all_dashboard_projects(data) if match_record(summary, p)[1]}
    unmatched_summary = [
        {
            "summaryProject": row["sourceProject"],
            "suites": row["suites"],
            "area": row["area"],
            "amountWan": row["amount"],
        }
        for key, row in summary.items()
        if key not in used_summary_keys and row["suites"]
    ]
    diff = {
        "month": MONTH,
        "dashboardProjects": len(data["projects"]),
        "matchedSummaryProjects": matched_projects,
        "addedProjects": added_projects,
        "summaryRecords": len(summary),
        "detailProjects": len(details),
        "diffRows": diff_rows,
        "unmatchedDashboard": unmatched_dashboard,
        "unmatchedSummaryWithSuites": unmatched_summary,
        **detail_extra,
        "dashboardJuneTotal": data["monthlyTotals"][-1],
        "detailJsSummary": detail_payload["summary"],
    }
    args.diff_json.write_text(json.dumps(diff, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "month": MONTH,
        "matchedSummaryProjects": matched_projects,
        "addedProjects": len(added_projects),
        "diffRows": len(diff_rows),
        "unmatchedDashboard": len(unmatched_dashboard),
        "unmatchedSummaryWithSuites": len(unmatched_summary),
        "unmatchedDetail": len(detail_extra["unmatchedDetail"]),
        "dashboardJuneTotal": data["monthlyTotals"][-1],
        "detailJsSummary": detail_payload["summary"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

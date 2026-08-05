#!/usr/bin/env python3
"""Update dashboard monthly metrics from a folder of CRIC detail workbooks."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATA_RE = re.compile(
    r"const DATA = (.*?);\n(const PROJECT_METADATA_OVERRIDES|const LAUNCH_OFFICIAL_INVENTORY_OVERRIDES|const DEFAULT_PERIODS)",
    re.S,
)
PROJECT_ALIASES_RE = re.compile(r"const PROJECT_NAME_ALIASES = \{(.*?)\};", re.S)
ZJW_NEW_LAUNCH_RE = re.compile(r"const ZJW_OFFICIAL_NEW_LAUNCH_PROJECTS = (\[.*?\]);", re.S)
SCRIPT_TAG_RE = re.compile(r'(<script src="june_transaction_details\.js[^"]*"></script>)')
ALLOWED_PROPERTY_TYPES = {"普通住宅", "别墅"}


def clean(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    return "" if value is None else str(value).strip()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).lower()
    replacements = {
        "雲": "云",
        "澐": "云",
        "樹": "树",
        "華": "华",
        "鳴": "鸣",
        "號": "号",
        "鄕": "乡",
        "臺": "台",
        "灣": "湾",
        "萬": "万",
        "叁": "三",
        "贰": "二",
        "壹": "一",
        "玖": "九",
        "·": "",
        "•": "",
        ".": "",
        "。": "",
        "\u05b7": "",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[\s\-—_/\\()（）【】\[\]《》“”\"'：:；;，,]+", "", text)


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def rounded(value: float, digits: int = 2) -> int | float:
    value = round(float(value), digits)
    if abs(value - round(value)) < 0.000001:
        return int(round(value))
    return value


def calc_price(amount_wan: float, area: float) -> int:
    return int(round(amount_wan * 10000 / area)) if area else 0


def month_full_label(month: str) -> str:
    match = re.match(r"(\d{2})年(\d{1,2})月", month)
    return f"{2000 + int(match.group(1))}年{int(match.group(2))}月" if match else month


def month_code(month: str) -> str:
    match = re.match(r"\d{2}年(\d{1,2})月", month)
    code_map = {
        1: "jan", 2: "feb", 3: "mar", 4: "apr", 5: "may", 6: "jun",
        7: "jul", 8: "aug", 9: "sep", 10: "oct", 11: "nov", 12: "dec",
    }
    return code_map.get(int(match.group(1)), "month") if match else "month"


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    text = clean(value)
    if " " in text:
        text = text.split(" ", 1)[0]
    return text.replace("-", "/")


def label_from_date(value: Any) -> str:
    if isinstance(value, datetime):
        return f"{value.year % 100:02d}年{value.month}月"
    text = clean(value)
    match = re.match(r"(\d{4})[-/](\d{1,2})", text)
    if not match:
        return ""
    return f"{int(match.group(1)) % 100:02d}年{int(match.group(2))}月"


def row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}


def find_header_row(rows: list[tuple[Any, ...]], required: str = "成交时间") -> int:
    for index, row in enumerate(rows):
        if required in [clean(value) for value in row]:
            return index
    raise ValueError(f"Cannot find header row with {required}")


def detail_paths(folder: Path) -> list[Path]:
    return [
        path for path in sorted(folder.glob("*.xlsx"))
        if path.is_file() and not path.name.startswith(("~", ".~"))
        and "汇总" not in path.stem
    ]


def detail_row(row: dict[str, Any], source_file: str) -> dict[str, Any]:
    total_wan = number(row.get("成交总价(万元)")) or number(row.get("成交总价(元)")) / 10000
    return {
        "date": date_text(row.get("成交时间")),
        "permit": clean(row.get("预售证号")),
        "sourceProject": clean(row.get("项目名称")),
        "building": clean(row.get("楼栋名称")),
        "unit": clean(row.get("单元号")),
        "room": clean(row.get("室号")),
        "propertyType": clean(row.get("物业类型")),
        "layout": clean(row.get("房型")),
        "area": rounded(number(row.get("成交面积(㎡)")), 2),
        "unitPrice": rounded(number(row.get("成交单价(元/㎡)")), 0),
        "totalWan": rounded(total_wan, 2),
        "_sourceFile": source_file,
    }


def row_signature(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row.get("date"),
        normalize_name(row.get("sourceProject")),
        row.get("permit"),
        row.get("building"),
        row.get("unit"),
        row.get("room"),
        row.get("area"),
        row.get("totalWan"),
    )


def load_details(folder: Path, month: str) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    stats = {
        "files": 0,
        "rawRows": 0,
        "usedRows": 0,
        "skippedRows": 0,
        "duplicateRows": 0,
        "excludedRows": 0,
        "propertyTypes": defaultdict(int),
    }
    seen: set[tuple[Any, ...]] = set()

    for path in detail_paths(folder):
        stats["files"] += 1
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            header_index = find_header_row(rows)
            headers = [clean(value) for value in rows[header_index]]
            for values in rows[header_index + 1:]:
                row = row_dict(headers, values)
                project_name = clean(row.get("项目名称"))
                trade_date = clean(row.get("成交时间"))
                if not project_name or not trade_date:
                    stats["skippedRows"] += 1
                    continue
                stats["rawRows"] += 1
                if label_from_date(row.get("成交时间")) != month:
                    stats["skippedRows"] += 1
                    continue
                property_type = clean(row.get("物业类型"))
                stats["propertyTypes"][property_type or "未填写"] += 1
                if property_type and property_type not in ALLOWED_PROPERTY_TYPES:
                    stats["excludedRows"] += 1
                    continue
                item = detail_row(row, path.name)
                signature = row_signature(item)
                if signature in seen:
                    stats["duplicateRows"] += 1
                    continue
                seen.add(signature)
                key = normalize_name(project_name)
                group = grouped.setdefault(
                    key,
                    {
                        "sourceProject": project_name,
                        "rows": [],
                        "sourceFiles": set(),
                    },
                )
                group["rows"].append(item)
                group["sourceFiles"].add(path.name)
                stats["usedRows"] += 1
        finally:
            wb.close()

    for group in grouped.values():
        rows = group["rows"]
        rows.sort(key=lambda item: item["date"], reverse=True)
        area = sum(number(row["area"]) for row in rows)
        amount = sum(number(row["totalWan"]) for row in rows)
        group["summary"] = {
            "suites": len(rows),
            "area": rounded(area, 2),
            "amountWan": rounded(amount, 2),
            "avgPrice": calc_price(amount, area),
        }
        group["sourceFiles"] = sorted(group["sourceFiles"])
        for row in rows:
            row.pop("_sourceFile", None)

    stats["propertyTypes"] = dict(sorted(stats["propertyTypes"].items()))
    return grouped, stats


def refresh_detail_summary(group: dict[str, Any]) -> None:
    rows = group["rows"]
    rows.sort(key=lambda item: item["date"], reverse=True)
    area = sum(number(row["area"]) for row in rows)
    amount = sum(number(row["totalWan"]) for row in rows)
    group["summary"] = {
        "suites": len(rows),
        "area": rounded(area, 2),
        "amountWan": rounded(amount, 2),
        "avgPrice": calc_price(amount, area),
    }
    group["sourceFiles"] = sorted(set(group.get("sourceFiles") or []))


def merge_details_by_alias_groups(
    details: dict[str, dict[str, Any]],
    alias_groups: list[list[str]],
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    for group in alias_groups:
        normalized_names = [normalize_name(name) for name in group if normalize_name(name)]
        present_keys = [key for key in normalized_names if key in details]
        if len(present_keys) < 2:
            continue
        target_key = normalized_names[0] if normalized_names[0] in details else present_keys[0]
        target = details[target_key]
        consumed: list[str] = []
        for key in present_keys:
            if key == target_key or key not in details:
                continue
            source = details.pop(key)
            target["rows"].extend(source.get("rows") or [])
            target["sourceFiles"] = [
                *(target.get("sourceFiles") or []),
                *(source.get("sourceFiles") or []),
            ]
            consumed.append(source.get("sourceProject") or key)
        if consumed:
            refresh_detail_summary(target)
            merged.append({
                "canonical": group[0],
                "targetDetailProject": target.get("sourceProject"),
                "mergedDetailProjects": consumed,
                "suites": target["summary"]["suites"],
            })
    return merged


def parse_alias_groups(html: str) -> list[list[str]]:
    match = PROJECT_ALIASES_RE.search(html)
    if not match:
        return []
    groups: list[list[str]] = []
    for canonical, aliases_text in re.findall(r'"([^"]+)"\s*:\s*\[([^\]]*)\]', match.group(1)):
        aliases = re.findall(r'"([^"]+)"', aliases_text)
        groups.append([canonical, *aliases])
    return groups


def split_names(value: Any) -> list[str]:
    return [part.strip() for part in re.split(r"[；;、/\n]+", clean(value)) if part.strip()]


def project_candidate_names(project: dict[str, Any], alias_groups: list[list[str]]) -> list[str]:
    fields = [
        "project",
        "matchedName",
        "summaryRecordName",
        "officialProjectName",
        "cricProjectName",
        "janAprMatchedName",
        "junMatchedName",
        "junCricProjectName",
        "julMatchedName",
        "julCricProjectName",
    ]
    names: list[str] = []
    for field in fields:
        names.extend(split_names(project.get(field)))
    normalized = {normalize_name(name) for name in names if normalize_name(name)}
    for group in alias_groups:
        if any(normalize_name(name) in normalized for name in group):
            names.extend(group)
    output: list[str] = []
    for name in names:
        name = clean(name)
        if name and name not in output:
            output.append(name)
    return output


def all_dashboard_projects(data: dict[str, Any]) -> list[dict[str, Any]]:
    projects: list[dict[str, Any]] = []
    seen: set[int] = set()
    for collection in ("projects", "launchProjects"):
        for project in data.get(collection, []):
            marker = id(project)
            if marker not in seen:
                seen.add(marker)
                projects.append(project)
    return projects


def parse_official_new_launch_items(html: str) -> list[dict[str, Any]]:
    match = ZJW_NEW_LAUNCH_RE.search(html)
    if not match:
        return []
    return json.loads(match.group(1), strict=False)


def official_new_launch_records(item: dict[str, Any]) -> list[dict[str, str]]:
    permits = item.get("permits") or []
    issue_dates = item.get("issueDates") or []
    return [
        {
            "name": clean(item.get("officialProjectName")),
            "permit": clean(permit),
            "date": clean(issue_dates[index] if index < len(issue_dates) else (issue_dates[0] if issue_dates else "")),
        }
        for index, permit in enumerate(permits)
    ]


def official_new_launch_project_id(item: dict[str, Any], index: int) -> str:
    first_permit = clean((item.get("permits") or [""])[0])
    match = re.search(r"\d{4}\)?\d+", first_permit)
    return f"zjw-launch-{match.group(0) if match else normalize_name(item.get('officialProjectName')) or index}"


def official_new_launch_residential_total(item: dict[str, Any]) -> int:
    return int(
        number(item.get("residentialTotal"))
        or number(item.get("approvedResidentialSuites"))
        or number(item.get("approvedTotalSuites"))
        or 0
    )


def project_presale_match_names(project: dict[str, Any], alias_groups: list[list[str]]) -> list[str]:
    return [normalize_name(name) for name in project_candidate_names(project, alias_groups) if len(normalize_name(name)) >= 3]


def presale_record_matches_project(record: dict[str, Any], project: dict[str, Any], alias_groups: list[list[str]]) -> bool:
    record_name = normalize_name(record.get("name"))
    project_names = project_presale_match_names(project, alias_groups)
    permit_text = clean(project.get("summaryPresalePermit"))
    if record.get("permit") and clean(record.get("permit")) in permit_text:
        return True
    return bool(
        record_name
        and any(
            name == record_name
            or (len(name) >= 4 and len(record_name) >= 4 and (name in record_name or record_name in name))
            for name in project_names
        )
    )


def build_official_new_launch_project(
    item: dict[str, Any],
    index: int,
    months: list[str],
    display_name: str,
    detail_source_name: str,
) -> dict[str, Any]:
    records = official_new_launch_records(item)
    issue_dates = sorted({record["date"] for record in records if record.get("date")})
    official_name = clean(item.get("officialProjectName"))
    monthly = {month: {"suites": 0, "area": 0, "price": 0, "amount": 0} for month in months}
    detail_urls = item.get("detailUrls") or []
    return {
        "id": official_new_launch_project_id(item, index),
        "group": item.get("group", ""),
        "plate": item.get("plate", ""),
        "project": display_name or official_name,
        "landDate": item.get("landDate", ""),
        "status": "新取证",
        "x": 50,
        "y": 50,
        "lat": number(item.get("lat")),
        "lng": number(item.get("lng")),
        "monthly": monthly,
        "source": "zjwNewLaunch",
        "suites34": 0,
        "area34": 0,
        "amount34": 0,
        "price4": 0,
        "amountAll": 0,
        "suitesAll": 0,
        "district": item.get("district", ""),
        "address": item.get("address", ""),
        "coordSource": "住建委坐落位置 + 业务板块近似定位",
        "coordSourceUrl": detail_urls[0] if detail_urls else "",
        "coordConfidence": item.get("coordConfidence", "低"),
        "coordSystem": "GCJ-02近似板块中心",
        "matchedName": detail_source_name or display_name or official_name,
        "summaryRecordName": official_name,
        "summaryPresalePermit": " / ".join(clean(permit) for permit in (item.get("permits") or []) if clean(permit)),
        "summaryDeveloper": item.get("developer", "住建委项目详情页"),
        "cricProjectName": detail_source_name or display_name,
        "officialProjectName": official_name,
        "officialResidentialTotal": official_new_launch_residential_total(item),
        "officialUnsignedSuites": None,
        "officialAvailableSuites": None,
        "officialUnsignedBlueSuites": 0,
        "officialBookedSuites": None,
        "officialContractSignedSuites": None,
        "officialFilingSuites": None,
        "officialSignedStatsSuites": None,
        "officialSignedStatsArea": None,
        "officialSignedStatsAvgPrice": None,
        "officialSignedSuites": None,
        "officialDetailSignedSuites": None,
        "officialInventoryEvidenceUrl": "\n".join(detail_urls),
        "officialInventoryFetchedAt": "",
        "officialInventoryMatchStatus": "住建委新发预售证项目；飞书映射表补充克而瑞案名",
        "officialInventoryTotalAuditNote": item.get("inventoryNote", ""),
        "approvedTotalSuites": item.get("approvedTotalSuites"),
        "presaleIssueRecords": records,
        "presaleIssueDates": issue_dates,
    }


def append_matching_official_new_launch_projects(
    data: dict[str, Any],
    html: str,
    details: dict[str, dict[str, Any]],
    alias_groups: list[list[str]],
) -> list[dict[str, Any]]:
    appended: list[dict[str, Any]] = []
    existing = all_dashboard_projects(data)
    launch_projects = data.setdefault("launchProjects", [])
    for index, item in enumerate(parse_official_new_launch_items(html)):
        if official_new_launch_residential_total(item) <= 0:
            continue
        records = official_new_launch_records(item)
        if any(any(presale_record_matches_project(record, project, alias_groups) for record in records) for project in existing):
            continue
        probe = build_official_new_launch_project(item, index, data["months"], clean(item.get("officialProjectName")), "")
        detail_key, detail, matched_by = match_detail(details, probe, alias_groups)
        if not detail:
            continue
        project = build_official_new_launch_project(
            item,
            index,
            data["months"],
            detail.get("sourceProject") or matched_by or clean(item.get("officialProjectName")),
            detail.get("sourceProject") or matched_by,
        )
        launch_projects.append(project)
        existing.append(project)
        appended.append({
            "project": project["project"],
            "officialProjectName": project["officialProjectName"],
            "matchedBy": matched_by,
            "detailProject": detail.get("sourceProject"),
            "suites": detail["summary"]["suites"],
        })
    return appended


def match_detail(details: dict[str, Any], project: dict[str, Any], alias_groups: list[list[str]]) -> tuple[str, dict[str, Any] | None, str]:
    for name in project_candidate_names(project, alias_groups):
        key = normalize_name(name)
        if key in details:
            return key, details[key], name
    return "", None, ""


def recalc_project(project: dict[str, Any], months: list[str]) -> None:
    recent_months = months[-2:]
    monthly = project.get("monthly", {})
    project["suites34"] = rounded(sum(number(monthly.get(month, {}).get("suites")) for month in recent_months))
    project["area34"] = rounded(sum(number(monthly.get(month, {}).get("area")) for month in recent_months), 2)
    project["amount34"] = rounded(sum(number(monthly.get(month, {}).get("amount")) for month in recent_months), 4)
    latest = monthly.get(recent_months[-1], {}) if recent_months else {}
    project["price4"] = latest.get("price", 0)
    project["suitesAll"] = rounded(sum(number(metric.get("suites")) for metric in monthly.values()))
    project["amountAll"] = rounded(sum(number(metric.get("amount")) for metric in monthly.values()), 4)


def aggregate_month(projects: list[dict[str, Any]], month: str) -> dict[str, Any]:
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects)
    return {
        "month": month,
        "suites": rounded(suites),
        "area": rounded(area, 2),
        "amount": rounded(amount, 4),
        "price": calc_price(amount, area),
    }


def refresh_aggregates(data: dict[str, Any]) -> None:
    months = data["months"]
    for project in all_dashboard_projects(data):
        recalc_project(project, months)

    projects = data.get("projects", [])
    recent = months[-2:]
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects for month in recent)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects for month in recent)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects for month in recent)
    data["totals"] = {
        "projects": len(projects),
        "plates": len({project.get("plate") for project in projects}),
        "active": sum(project.get("status") == "在售" for project in projects),
        "suites34": rounded(suites),
        "amount34": rounded(amount, 4),
        "area34": rounded(area, 2),
        "avgPrice34": calc_price(amount, area),
    }
    data["monthlyTotals"] = [aggregate_month(projects, month) for month in months]

    group_totals = []
    for group in dict.fromkeys(project.get("group") for project in projects):
        group_projects = [project for project in projects if project.get("group") == group]
        group_totals.append(
            {
                "group": group,
                "projects": len(group_projects),
                "suites34": rounded(sum(number(project.get("suites34")) for project in group_projects)),
                "amount34": rounded(sum(number(project.get("amount34")) for project in group_projects), 4),
                "color": data.get("colors", {}).get(group, "#8ea0b8"),
            }
        )
    data["groupTotals"] = group_totals
    data["colors"] = {row["group"]: row["color"] for row in group_totals}


def build_detail_payload(
    data: dict[str, Any],
    details: dict[str, dict[str, Any]],
    alias_groups: list[list[str]],
    month: str,
    detail_dir: Path,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    projects: dict[str, Any] = {}
    aliases: dict[str, str] = {}
    used_keys: set[str] = set()

    for project in all_dashboard_projects(data):
        detail_key, detail, matched_by = match_detail(details, project, alias_groups)
        if not detail:
            continue
        primary = normalize_name(project.get("project"))
        projects[primary] = {
            "projectName": project.get("project", ""),
            "rawProjectName": project.get("project", ""),
            "cricProjectName": detail.get("sourceProject", ""),
            "matchedProjectName": matched_by or project.get("project", ""),
            "plate": project.get("plate", ""),
            "group": project.get("group", ""),
            "rows": detail.get("rows", []),
            "summary": detail.get("summary", {}),
        }
        used_keys.add(detail_key)
        for alias in [*project_candidate_names(project, alias_groups), detail.get("sourceProject")]:
            key = normalize_name(alias)
            if key and key != primary:
                aliases[key] = primary

    unmatched = [
        {
            "detailProject": detail["sourceProject"],
            "suites": detail["summary"]["suites"],
            "area": detail["summary"]["area"],
            "amountWan": detail["summary"]["amountWan"],
            "sourceFiles": detail["sourceFiles"],
        }
        for key, detail in details.items()
        if key not in used_keys and detail["summary"]["suites"]
    ]
    payload = {
        "source": f"{detail_dir.name}（排除汇总表）",
        "sheet": "CRIC-北京-项目详情.交易.项目累计.成交明细",
        "scope": "普通住宅/别墅，已排除车库/车位",
        "month": month_full_label(month),
        "projects": projects,
        "aliases": aliases,
        "summary": {
            "projects": len(projects),
            "rows": sum(len(project["rows"]) for project in projects.values()),
            "excludedParkingRows": 0,
        },
    }
    return payload, unmatched


def write_detail_js(path: Path, global_name: str, month: str, payload: dict[str, Any]) -> None:
    merge_js = f"""
if (window.TRANSACTION_DETAILS) {{
  window.TRANSACTION_DETAILS.months = window.TRANSACTION_DETAILS.months || {{}};
  window.TRANSACTION_DETAILS.months["{month}"] = window.{global_name};
  window.TRANSACTION_DETAILS.summary = window.TRANSACTION_DETAILS.summary || {{}};
  const monthRows = monthData => Object.values(monthData.projects || {{}}).reduce((sum, project) => sum + ((project.rows || []).length), 0);
  const months = Object.values(window.TRANSACTION_DETAILS.months || {{}});
  window.TRANSACTION_DETAILS.summary.months = Object.keys(window.TRANSACTION_DETAILS.months).length;
  window.TRANSACTION_DETAILS.summary.projects = months.reduce((sum, monthData) => sum + Object.keys(monthData.projects || {{}}).length, 0);
  window.TRANSACTION_DETAILS.summary.rows = months.reduce((sum, monthData) => sum + monthRows(monthData), 0);
}}
""".strip()
    path.write_text(
        f"window.{global_name} = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
        + merge_js
        + "\n",
        encoding="utf-8",
    )


def update_html(
    html: str,
    data: dict[str, Any],
    data_replacement_suffix: str,
    month: str,
    detail_js_name: str,
) -> str:
    replacement = "const DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n" + data_replacement_suffix
    html = DATA_RE.sub(lambda _: replacement, html, count=1)
    html = re.sub(
        r'(<div class="cutoff"><i>▣</i><span>数据截至：</span><b>).*?(</b></div>)',
        rf"\g<1>{month_full_label(month)}\g<2>",
        html,
        count=1,
    )
    html = re.sub(
        r"const periodOptions = .*?;",
        "const periodOptions = "
        + json.dumps([[value, value] for value in reversed(data["months"])], ensure_ascii=False)
        + ";",
        html,
        count=1,
    )
    html = re.sub(
        r"const recentLaunchMonths = \[[^\]]*\]",
        "const recentLaunchMonths = " + json.dumps(data["months"][-2:], ensure_ascii=False),
        html,
        count=1,
    )
    if detail_js_name not in html:
        html = SCRIPT_TAG_RE.sub(rf'\1\n  <script src="{detail_js_name}?v=20260708"></script>', html, count=1)
    return html


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--html", type=Path, default=Path("index.html"))
    parser.add_argument("--detail-dir", type=Path, required=True)
    parser.add_argument("--month", required=True)
    parser.add_argument("--detail-js", type=Path, default=Path("july_transaction_details.js"))
    parser.add_argument("--diff-json", type=Path, default=Path("month_detail_update_diff.json"))
    args = parser.parse_args()

    html = args.html.read_text(encoding="utf-8")
    match = DATA_RE.search(html)
    if not match:
        raise RuntimeError("DATA block not found")
    data = json.loads(match.group(1), strict=False)
    alias_groups = parse_alias_groups(html)
    details, stats = load_details(args.detail_dir, args.month)
    alias_detail_merges = merge_details_by_alias_groups(details, alias_groups)

    if args.month not in data["months"]:
        data["months"].append(args.month)
    appended_launch_projects = append_matching_official_new_launch_projects(data, html, details, alias_groups)

    code = month_code(args.month)
    matched: list[dict[str, Any]] = []
    zeroed = 0
    for project in all_dashboard_projects(data):
        monthly = project.setdefault("monthly", {})
        for month in data["months"]:
            monthly.setdefault(month, {"suites": 0, "area": 0, "price": 0, "amount": 0})
        detail_key, detail, matched_by = match_detail(details, project, alias_groups)
        if detail:
            summary = detail["summary"]
            monthly[args.month] = {
                "suites": summary["suites"],
                "area": summary["area"],
                "price": summary["avgPrice"],
                "amount": summary["amountWan"],
            }
            project[f"{code}DataSource"] = "克尔瑞"
            project[f"{code}MatchedName"] = matched_by
            project[f"{code}CricProjectName"] = detail["sourceProject"]
            matched.append({
                "project": project.get("project"),
                "detailProject": detail["sourceProject"],
                "matchedBy": matched_by,
                "suites": summary["suites"],
            })
        else:
            monthly[args.month] = {"suites": 0, "area": 0, "price": 0, "amount": 0}
            zeroed += 1

    refresh_aggregates(data)
    policy = data.setdefault("sourcePolicy", {})
    policy[month_full_label(args.month)] = "克尔瑞项目成交明细逐套聚合（普通住宅/别墅，不读取汇总表）"
    policy[f"{args.month}成交明细"] = "克尔瑞项目成交明细（普通住宅/别墅，不读取汇总表）"

    payload, unmatched_detail = build_detail_payload(data, details, alias_groups, args.month, args.detail_dir)
    global_name = f"{code.upper()}_TRANSACTION_DETAILS"
    write_detail_js(args.detail_js, global_name, args.month, payload)

    updated_html = update_html(html, data, match.group(2), args.month, args.detail_js.name)
    args.html.write_text(updated_html, encoding="utf-8")

    diff = {
        "month": args.month,
        "detailStats": stats,
        "appendedLaunchProjects": appended_launch_projects,
        "aliasDetailMerges": alias_detail_merges,
        "matchedProjects": matched,
        "zeroedProjects": zeroed,
        "unmatchedDetail": unmatched_detail,
        "dashboardMonthTotal": next((row for row in data["monthlyTotals"] if row["month"] == args.month), None),
        "detailJsSummary": payload["summary"],
    }
    args.diff_json.write_text(json.dumps(diff, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "month": args.month,
        "matchedProjects": len(matched),
        "appendedLaunchProjects": len(appended_launch_projects),
        "aliasDetailMerges": len(alias_detail_merges),
        "unmatchedDetail": len(unmatched_detail),
        "detailStats": stats,
        "dashboardMonthTotal": diff["dashboardMonthTotal"],
        "detailJsSummary": payload["summary"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

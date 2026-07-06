#!/usr/bin/env python3
"""Build supplemental transaction details for dashboard launchProjects."""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATA_RE = re.compile(
    r"const DATA = (.*?);\n(?:const LAUNCH_OFFICIAL_INVENTORY_OVERRIDES|const DEFAULT_PERIODS)",
    re.S,
)
DETAIL_SCOPE = "普通住宅/别墅，已排除车库/车位"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).lower()
    replacements = {
        "雲": "云",
        "澐": "云",
        "樹": "树",
        "華": "华",
        "鳴": "鸣",
        "號": "号",
        "鄕": "乡",
        "臺": "台",
        "灣": "湾",
        "叁": "三",
        "贰": "二",
        "壹": "一",
        "玖": "九",
        "·": "",
        "•": "",
        ".": "",
        "。": "",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[\s\-—_/\\()（）【】\[\]《》“”\"'：:；;，,]+", "", text)


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def rounded(value: float, digits: int = 2) -> int | float:
    value = round(float(value), digits)
    if abs(value - round(value)) < 0.000001:
        return int(round(value))
    return value


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    text = clean(value)
    if " " in text:
        text = text.split(" ", 1)[0]
    return text.replace("-", "/")


def month_label(value: Any) -> str:
    if isinstance(value, datetime):
        return f"{value.year % 100:02d}年{value.month}月"
    text = clean(value)
    match = re.match(r"(\d{4})[-/](\d{1,2})", text)
    if not match:
        return ""
    return f"{int(match.group(1)) % 100:02d}年{int(match.group(2))}月"


def load_dashboard(path: Path) -> dict[str, Any]:
    match = DATA_RE.search(path.read_text(encoding="utf-8"))
    if not match:
        raise ValueError(f"DATA block not found: {path}")
    return json.loads(match.group(1), strict=False)


def candidate_names(project: dict[str, Any]) -> list[str]:
    fields = [
        "project",
        "matchedName",
        "cricProjectName",
        "janAprMatchedName",
        "junMatchedName",
        "junCricProjectName",
        "summaryRecordName",
        "officialProjectName",
    ]
    manual = {
        "北投栖澐湾": ["北投·云帆汀澜", "北投云帆汀澜"],
        "中海玖樹满和": ["中海·九树满和", "中海九树满和"],
        "中建方程国贤府": ["方程国贤府", "中建·方程国贤府"],
        "未来城星寰时代": ["未来城·星寰时代"],
    }
    names: list[str] = []
    for field in fields:
        value = clean(project.get(field))
        if value:
            names.extend(re.split(r"[；;、/\n]+", value))
    names.extend(manual.get(clean(project.get("project")), []))
    output: list[str] = []
    for name in names:
        name = clean(name)
        if name and name not in output:
            output.append(name)
    return output


def build_alias_index(projects: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    aliases: dict[str, dict[str, Any]] = {}
    for project in projects:
        for name in candidate_names(project):
            key = normalize_name(name)
            if key:
                aliases[key] = project
    return aliases


def row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}


def detail_row(row: dict[str, Any]) -> dict[str, Any]:
    total_wan = number(row.get("trade_amount")) / 10000
    return {
        "date": date_text(row.get("trade_day")),
        "permit": clean(row.get("pre_permit")),
        "sourceProject": clean(row.get("project_name")),
        "building": clean(row.get("building_name")),
        "unit": "",
        "room": clean(row.get("room_number")),
        "propertyType": clean(row.get("property_type")),
        "layout": clean(row.get("layout_type")),
        "area": rounded(number(row.get("trade_area")), 2),
        "unitPrice": rounded(number(row.get("trade_price")), 0),
        "totalWan": rounded(total_wan, 2),
    }


def add_aliases(alias_map: dict[str, str], primary: str, project: dict[str, Any], source_project: str) -> None:
    for alias in [*candidate_names(project), source_project]:
        key = normalize_name(alias)
        if key and key != primary:
            alias_map.setdefault(key, primary)


def recompute_summary(group: dict[str, Any]) -> None:
    rows = group.get("rows", [])
    area = sum(number(row.get("area")) for row in rows)
    amount = sum(number(row.get("totalWan")) for row in rows)
    group["summary"] = {
        "suites": len(rows),
        "area": rounded(area, 2),
        "amountWan": rounded(amount, 2),
        "avgPrice": int(round(amount * 10000 / area)) if area else 0,
    }


def build_payload(data: dict[str, Any], source: Path) -> dict[str, Any]:
    launch_projects = data.get("launchProjects", [])
    alias_index = build_alias_index(launch_projects)
    months: dict[str, Any] = {}
    property_counts: dict[str, int] = defaultdict(int)
    skipped = 0
    excluded = 0

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        grouped: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
        aliases_by_month: dict[str, dict[str, str]] = defaultdict(dict)
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = row_dict(headers, values)
            if clean(row.get("city_name")) != "北京":
                continue
            project = alias_index.get(normalize_name(row.get("project_name")))
            if not project:
                continue
            month = month_label(row.get("trade_day"))
            if month not in data.get("months", []):
                skipped += 1
                continue
            property_type = clean(row.get("property_type")) or "未填写"
            property_counts[property_type] += 1
            if property_type not in {"普通住宅", "别墅"}:
                excluded += 1
                continue
            primary = normalize_name(project.get("project"))
            month_projects = grouped[month]
            group = month_projects.get(primary)
            if not group:
                group = {
                    "projectName": project.get("project", ""),
                    "rawProjectName": project.get("project", ""),
                    "cricProjectName": clean(row.get("project_name")),
                    "matchedProjectName": project.get("junMatchedName") or project.get("cricProjectName") or project.get("project", ""),
                    "plate": project.get("plate", ""),
                    "group": project.get("group", ""),
                    "rows": [],
                    "summary": {"suites": 0, "area": 0, "amountWan": 0, "avgPrice": 0},
                }
                month_projects[primary] = group
            group["rows"].append(detail_row(row))
            add_aliases(aliases_by_month[month], primary, project, clean(row.get("project_name")))

        for month, month_projects in sorted(grouped.items()):
            projects: dict[str, Any] = {}
            row_count = 0
            for key, group in sorted(month_projects.items(), key=lambda item: item[1]["projectName"]):
                group["rows"].sort(key=lambda item: item["date"], reverse=True)
                recompute_summary(group)
                row_count += len(group["rows"])
                projects[key] = group
            months[month] = {
                "projects": projects,
                "aliases": aliases_by_month.get(month, {}),
                "summary": {"projects": len(projects), "rows": row_count},
            }
    finally:
        wb.close()

    return {
        "source": source.name,
        "sheet": "Sheet1",
        "scope": DETAIL_SCOPE,
        "months": months,
        "summary": {
            "months": len(months),
            "projects": sum(month["summary"]["projects"] for month in months.values()),
            "rows": sum(month["summary"]["rows"] for month in months.values()),
            "skippedRows": skipped,
            "excludedParkingRows": excluded,
            "propertyTypes": dict(sorted(property_counts.items())),
        },
    }


def write_js(output: Path, payload: dict[str, Any]) -> None:
    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    output.write_text(
        "window.NEW_LAUNCH_TRANSACTION_DETAILS = "
        + body
        + ";\n"
        + r"""
(function(){
  const incoming = window.NEW_LAUNCH_TRANSACTION_DETAILS;
  if (!incoming) return;
  window.TRANSACTION_DETAILS = window.TRANSACTION_DETAILS || { months:{}, summary:{ months:0, projects:0, rows:0 } };
  window.TRANSACTION_DETAILS.months = window.TRANSACTION_DETAILS.months || {};
  const monthRows = monthData => Object.values(monthData.projects || {}).reduce((sum, project) => sum + ((project.rows || []).length), 0);
  for (const [month, monthData] of Object.entries(incoming.months || {})) {
    const target = window.TRANSACTION_DETAILS.months[month] || { projects:{}, aliases:{}, summary:{ projects:0, rows:0 } };
    target.projects = target.projects || {};
    for (const [projectKey, projectValue] of Object.entries(monthData.projects || {})) {
      if (!target.projects[projectKey]) target.projects[projectKey] = projectValue;
    }
    target.aliases = Object.assign({}, monthData.aliases || {}, target.aliases || {});
    target.summary = {
      projects: Object.keys(target.projects || {}).length,
      rows: monthRows(target)
    };
    window.TRANSACTION_DETAILS.months[month] = target;
  }
  const months = Object.values(window.TRANSACTION_DETAILS.months || {});
  window.TRANSACTION_DETAILS.summary = Object.assign({}, window.TRANSACTION_DETAILS.summary || {}, {
    months: months.length,
    projects: months.reduce((sum, month) => sum + Object.keys(month.projects || {}).length, 0),
    rows: months.reduce((sum, month) => sum + monthRows(month), 0)
  });
  window.MAY_TRANSACTION_DETAILS = window.TRANSACTION_DETAILS.months["26年5月"] || { projects:{}, aliases:{}, summary:{ projects:0, rows:0 } };
})();
""".lstrip(),
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--html", type=Path, default=Path("index.html"))
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--out", type=Path, default=Path("new_launch_transaction_details.js"))
    args = parser.parse_args()

    data = load_dashboard(args.html)
    payload = build_payload(data, args.source)
    write_js(args.out, payload)
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

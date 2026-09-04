#!/usr/bin/env python3
"""Patch August dashboard data from the mapped CRIC database export."""

from __future__ import annotations

import argparse
import copy
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import replace_transaction_history_from_xlsx as base


MONTH = "26年8月"
MONTH_FULL = "2026年8月"
DATA_RE = base.DATA_RE
OFFICIAL_RE = base.OFFICIAL_NEW_LAUNCH_RE
DETAIL_RE = re.compile(
    r"window\.TRANSACTION_DETAILS = (.*?);\nwindow\.MAY_TRANSACTION_DETAILS",
    re.S,
)

OFFICIAL_PROJECT_ENRICHMENT = {
    "元著嘉苑": {"group": "西南组团", "plate": "西红门", "displayName": "中建壹品·未来元著"},
    "世和承园": {"group": "东北组团", "plate": "太阳宫", "lat": 39.973, "lng": 116.447, "coordConfidence": "低"},
    "瑞璟佳苑": {"group": "西部组团", "plate": "四季青", "displayName": "保利熙瑞"},
    "海岄雅苑": {"group": "西部组团", "plate": "四季青", "displayName": "北京隅海岄"},
    "誉淙家园": {"group": "北部组团", "plate": "回龙观", "displayName": "国誉星城"},
    "青年万序家园": {"group": "东北组团", "plate": "南彩", "lat": 40.136, "lng": 116.728, "coordConfidence": "低"},
    "润樾雅苑": {"group": "北部组团", "plate": "北七家"},
    "和璟嘉苑": {"group": "CBD-副中心组团", "plate": "宋庄东", "displayName": "北投和璟"},
    "宸章院": {"group": "西南组团", "plate": "良乡"},
    "满茂文苑": {"group": "东北组团", "plate": "东坝", "lat": 39.949, "lng": 116.558, "coordConfidence": "低"},
    "朗月和风雅苑": {"group": "西南组团", "plate": "黄村", "displayName": "北京朗月和风", "lat": 39.731, "lng": 116.341, "coordConfidence": "低"},
}

CRIC_NAME_ALIASES = {
    "和璟嘉苑": ["北投和璟"],
    "朗月和风雅苑": ["北京朗月和风"],
    "元著嘉苑": ["中建壹品?未来元著", "中建壹品·未来元著"],
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> float:
    return base.number(value)


def rounded(value: float, digits: int = 2) -> int | float:
    return base.rounded(value, digits)


def calc_price(amount_wan: float, area: float) -> int:
    return int(round(amount_wan * 10000 / area)) if area else 0


def normalize_plate(value: Any) -> str:
    return clean(value).removesuffix("板块")


def load_dashboard(html_path: Path) -> tuple[str, dict[str, Any], str]:
    html = html_path.read_text(encoding="utf-8")
    match = DATA_RE.search(html)
    if not match:
        raise RuntimeError("DATA block not found")
    return html, json.loads(match.group(1), strict=False), match.group(2)


def load_transaction_details(path: Path) -> dict[str, Any]:
    match = DETAIL_RE.search(path.read_text(encoding="utf-8"))
    if not match:
        raise RuntimeError(f"TRANSACTION_DETAILS block not found: {path}")
    return json.loads(match.group(1), strict=False)


def write_transaction_details(path: Path, details: dict[str, Any]) -> None:
    base.write_transaction_js(path, details)


def recompute_transaction_summary(details: dict[str, Any]) -> None:
    months = details.get("months", {})
    property_types: Counter[str] = Counter()
    project_count = 0
    row_count = 0
    for month_data in months.values():
        projects = month_data.get("projects", {})
        project_count += len(projects)
        for project in projects.values():
            rows = project.get("rows", [])
            row_count += len(rows)
            for row in rows:
                property_types[clean(row.get("propertyType")) or "未填写"] += 1
    details["summary"] = {
        "months": len(months),
        "projects": project_count,
        "rows": row_count,
        "skippedRows": 0,
        "excludedParkingRows": 0,
        "propertyTypes": dict(sorted(property_types.items())),
    }


def load_cric_summary(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if "项目名称" in [clean(value) for value in row]:
                header_row = idx
                headers = [clean(value) for value in row]
                break
        if header_row is None:
            raise RuntimeError("CRIC summary header not found")
        index = {header: pos for pos, header in enumerate(headers)}
        rows: dict[str, dict[str, Any]] = {}
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = clean(values[index["项目名称"]])
            if not name:
                continue
            amount_wan = number(values[index["成交金额(元)"]]) / 10000
            area = number(values[index["成交面积(㎡)"]])
            row = {
                "sourceProject": name,
                "district": clean(values[index["区域"]]),
                "plate": normalize_plate(values[index["板块"]]),
                "developer": clean(values[index["开发商"]]),
                "supplyArea": rounded(number(values[index["供应面积(㎡)"]]), 2),
                "supplySuites": rounded(number(values[index["供应套数(套)"]])),
                "area": rounded(area, 2),
                "suites": rounded(number(values[index["成交套数(套)"]])),
                "price": rounded(number(values[index["成交均价(元/㎡)"]]), 0),
                "amount": rounded(amount_wan, 4),
                "landInfo": clean(values[index["土地信息"]]),
            }
            rows[base.normalize_name(name)] = row
        return rows
    finally:
        wb.close()


def project_candidate_names(project: dict[str, Any], alias_groups: list[list[str]]) -> list[str]:
    names: list[str] = []
    fields = [
        *base.PROJECT_FIELDS,
        "augMatchedName",
        "augCricProjectName",
    ]
    for field in fields:
        names.extend(base.split_names(project.get(field)))
    return base.expanded_names(names, alias_groups)


def find_summary_row(names: list[str], summary_rows: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    normalized = [base.normalize_name(name) for name in names if base.normalize_name(name)]
    for key in normalized:
        if key in summary_rows:
            return summary_rows[key]
    for key in normalized:
        if len(key) < 3:
            continue
        for summary_key, row in summary_rows.items():
            if key == summary_key or (len(summary_key) >= 3 and (key in summary_key or summary_key in key)):
                return row
    return None


def summarize_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return base.summarize_rows(rows)


def transaction_detail_house_key(row: dict[str, Any]) -> tuple[str, ...]:
    building = clean(row.get("building"))
    room = clean(row.get("room"))
    if building and room:
        return (
            clean(row.get("sourceProject")),
            clean(row.get("permit")),
            building,
            clean(row.get("unit")),
            room,
        )
    return (
        "row",
        clean(row.get("date")),
        clean(row.get("sourceProject")),
        building,
        clean(row.get("unit")),
        room,
        str(number(row.get("area"))),
        str(number(row.get("totalWan"))),
    )


def summarize_transaction_detail_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    latest_by_house: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        key = transaction_detail_house_key(row)
        date = clean(row.get("date"))
        existing = latest_by_house.get(key)
        if not existing or date >= clean(existing.get("date")):
            latest_by_house[key] = row
    unique_rows = list(latest_by_house.values())
    dates = [clean(row.get("date")) for row in unique_rows if clean(row.get("date"))]
    return {
        "suites": len(unique_rows),
        "recordRows": len(rows),
        "duplicateRows": len(rows) - len(unique_rows),
        "area": rounded(sum(number(row.get("area")) for row in unique_rows), 2),
        "amountWan": rounded(sum(number(row.get("totalWan")) for row in unique_rows), 4),
        "firstDate": min(dates) if dates else "",
        "lastDate": max(dates) if dates else "",
    }


def detail_rows_for_project(
    details: dict[str, Any],
    project: dict[str, Any],
    alias_groups: list[list[str]],
    months: list[str] | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    month_names = months or sorted(details.get("months", {}), key=base.month_tuple)
    for month in month_names:
        month_data = details.get("months", {}).get(month, {})
        projects = month_data.get("projects", {})
        aliases = month_data.get("aliases", {})
        candidates = [
            base.normalize_name(project.get("project")),
            *(base.normalize_name(name) for name in project_candidate_names(project, alias_groups)),
        ]
        for candidate in dict.fromkeys(name for name in candidates if name):
            detail_key = candidate
            if detail_key not in projects:
                detail_key = aliases.get(candidate, "")
            if detail_key in projects:
                rows.extend(projects[detail_key].get("rows", []))
                break
    return rows


def refresh_historical_sales_from_august_delta(
    data: dict[str, Any],
    old_details: dict[str, Any],
    new_details: dict[str, Any],
    alias_groups: list[list[str]],
    max_date: str,
) -> dict[str, Any]:
    updated = 0
    source_file = "项目成交明细.xlsx（25年1月-26年7月原历史累计） + 陈麒亦需求.xlsx（26年8月）"
    for project in base.all_projects(data):
        old_summary = summarize_transaction_detail_rows(
            detail_rows_for_project(old_details, project, alias_groups, [MONTH])
        )
        new_summary = summarize_transaction_detail_rows(
            detail_rows_for_project(new_details, project, alias_groups, [MONTH])
        )
        has_history = "historicalTransactionSoldSuites" in project
        if not has_history and not old_summary["recordRows"] and not new_summary["recordRows"]:
            continue

        all_detail_summary = summarize_transaction_detail_rows(
            detail_rows_for_project(new_details, project, alias_groups)
        )
        project["historicalTransactionSoldSuites"] = max(
            0,
            int(number(project.get("historicalTransactionSoldSuites")))
            + int(new_summary["suites"])
            - int(old_summary["suites"]),
        )
        project["historicalTransactionRecordRows"] = max(
            0,
            int(number(project.get("historicalTransactionRecordRows")))
            + int(new_summary["recordRows"])
            - int(old_summary["recordRows"]),
        )
        project["historicalTransactionDuplicateRows"] = max(
            0,
            int(number(project.get("historicalTransactionDuplicateRows")))
            + int(new_summary["duplicateRows"])
            - int(old_summary["duplicateRows"]),
        )
        project["historicalTransactionArea"] = rounded(
            number(project.get("historicalTransactionArea"))
            + number(new_summary["area"])
            - number(old_summary["area"]),
            2,
        )
        project["historicalTransactionAmountWan"] = rounded(
            number(project.get("historicalTransactionAmountWan"))
            + number(new_summary["amountWan"])
            - number(old_summary["amountWan"]),
            4,
        )
        first_date = clean(project.get("historicalTransactionStartDate"))
        if all_detail_summary["firstDate"] and (
            not first_date or all_detail_summary["firstDate"] < first_date
        ):
            first_date = all_detail_summary["firstDate"]
        project["historicalTransactionStartDate"] = first_date
        if all_detail_summary["lastDate"]:
            project["historicalTransactionEndDate"] = all_detail_summary["lastDate"]
        project["historicalTransactionCoverageStartDate"] = (
            clean(project.get("historicalTransactionCoverageStartDate"))
            or all_detail_summary["firstDate"]
        )
        project["historicalTransactionCoverageEndDate"] = max_date
        project["historicalTransactionSource"] = source_file
        project["historicalTransactionScope"] = "仅统计2025年1月至今普通住宅/别墅；同一房源多次成交按最新一笔去重计数"
        updated += 1
    return {"updatedProjects": updated}


def replace_august_dashboard_month(
    data: dict[str, Any],
    payload: dict[str, Any],
    alias_groups: list[list[str]],
) -> dict[str, Any]:
    month_payload = payload["months"].get(MONTH, {"projects": {}})
    payload_by_project = {
        key: group["summary"]
        for key, group in month_payload.get("projects", {}).items()
    }
    matched = []
    zeroed = 0
    for project in base.all_projects(data):
        primary = base.normalize_name(project.get("project"))
        monthly = project.setdefault("monthly", {})
        for month in data.get("months", []):
            monthly.setdefault(month, {"suites": 0, "area": 0, "price": 0, "amount": 0})
        summary = payload_by_project.get(primary)
        if summary:
            monthly[MONTH] = {
                "suites": summary["suites"],
                "area": summary["area"],
                "price": summary["avgPrice"],
                "amount": summary["amountWan"],
            }
            month_data = month_payload["projects"].get(primary, {})
            project["augDataSource"] = "克尔瑞"
            project["augMatchedName"] = month_data.get("matchedProjectName") or project.get("project", "")
            project["augCricProjectName"] = month_data.get("cricProjectName") or project.get("cricProjectName", "")
            matched.append({"project": project.get("project"), **monthly[MONTH]})
        else:
            monthly[MONTH] = {"suites": 0, "area": 0, "price": 0, "amount": 0}
            zeroed += 1

    data["sourcePolicy"][MONTH_FULL] = "克而瑞项目成交明细逐套聚合（普通住宅/别墅；本月增量替换）"
    data["sourcePolicy"][f"{MONTH}成交明细"] = "克而瑞项目成交明细（普通住宅/别墅；本月增量替换）"
    return {"matched": matched, "zeroed": zeroed}


def recalc_project(project: dict[str, Any], months: list[str]) -> None:
    monthly = project.setdefault("monthly", {})
    recent = months[-2:]
    project["suites34"] = rounded(sum(number(monthly.get(month, {}).get("suites")) for month in recent))
    project["area34"] = rounded(sum(number(monthly.get(month, {}).get("area")) for month in recent), 2)
    project["amount34"] = rounded(sum(number(monthly.get(month, {}).get("amount")) for month in recent), 4)
    project["price4"] = monthly.get(months[-1], {}).get("price", 0) if months else 0
    project["suitesAll"] = rounded(sum(number(metric.get("suites")) for metric in monthly.values()))
    project["amountAll"] = rounded(sum(number(metric.get("amount")) for metric in monthly.values()), 4)


def aggregate_month(projects: list[dict[str, Any]], month: str) -> dict[str, Any]:
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects)
    return {
        "month": month,
        "suites": rounded(suites),
        "area": rounded(area, 2),
        "amount": rounded(amount, 4),
        "price": calc_price(amount, area),
    }


def refresh_dashboard_aggregates(data: dict[str, Any]) -> None:
    months = data.get("months", [])
    for project in base.all_projects(data):
        for month in months:
            project.setdefault("monthly", {}).setdefault(month, {"suites": 0, "area": 0, "price": 0, "amount": 0})
        recalc_project(project, months)

    projects = data.get("projects", [])
    recent = months[-2:]
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects for month in recent)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects for month in recent)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects for month in recent)
    data["totals"] = {
        "projects": len(projects),
        "plates": len({project.get("plate") for project in projects}),
        "active": sum(project.get("status") == "在售" for project in projects),
        "suites34": rounded(suites),
        "amount34": rounded(amount, 4),
        "area34": rounded(area, 2),
        "avgPrice34": calc_price(amount, area),
    }
    data["monthlyTotals"] = [aggregate_month(projects, month) for month in months]

    group_totals = []
    for group in dict.fromkeys(project.get("group") for project in projects):
        group_projects = [project for project in projects if project.get("group") == group]
        group_totals.append({
            "group": group,
            "projects": len(group_projects),
            "suites34": rounded(sum(number(project.get("suites34")) for project in group_projects)),
            "amount34": rounded(sum(number(project.get("amount34")) for project in group_projects), 4),
            "color": data.get("colors", {}).get(group, "#8ea0b8"),
        })
    data["groupTotals"] = group_totals
    data["colors"] = {row["group"]: row["color"] for row in group_totals}


def summary_validation(
    grouped: dict[str, dict[str, list[dict[str, Any]]]],
    stats: dict[str, Any],
    matches: dict[str, dict[str, Any]],
    summary_rows: dict[str, dict[str, Any]],
    alias_groups: list[list[str]],
) -> dict[str, Any]:
    diffs = []
    unmatched = []
    for mapped_key, project in matches.items():
        rows = grouped.get(MONTH, {}).get(mapped_key, [])
        candidates = [
            *stats.get("sourceProjects", {}).get(mapped_key, []),
            stats.get("historicalMetrics", {}).get(mapped_key, {}).get("mappedName", ""),
            *project_candidate_names(project, alias_groups),
        ]
        summary = find_summary_row(candidates, summary_rows)
        detail_summary = summarize_rows(rows)
        if not summary:
            unmatched.append({
                "project": project.get("project"),
                "mappedKey": mapped_key,
                "detailSuites": detail_summary["suites"],
                "candidates": [name for name in candidates if clean(name)],
            })
            continue
        suite_diff = int(number(summary["suites"]) - number(detail_summary["suites"]))
        area_diff = round(number(summary["area"]) - number(detail_summary["area"]), 2)
        amount_diff = round(number(summary["amount"]) - number(detail_summary["amountWan"]), 2)
        if suite_diff or abs(area_diff) > 1 or abs(amount_diff) > 0.05:
            diffs.append({
                "project": project.get("project"),
                "summaryProject": summary["sourceProject"],
                "detailSuites": detail_summary["suites"],
                "summarySuites": summary["suites"],
                "suiteDiff": suite_diff,
                "areaDiff": area_diff,
                "amountDiffWan": amount_diff,
            })
    return {
        "checkedProjects": len(matches),
        "unmatchedSummary": unmatched,
        "diffs": diffs,
    }


def plate_group_lookup(data: dict[str, Any]) -> dict[str, str]:
    groups: dict[str, Counter[str]] = defaultdict(Counter)
    for project in base.all_projects(data):
        plate = clean(project.get("plate"))
        group = clean(project.get("group"))
        if plate and group:
            groups[plate][group] += 1
    return {plate: counter.most_common(1)[0][0] for plate, counter in groups.items()}


def cric_summary_for_official(item: dict[str, Any], summary_rows: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    names = [
        clean(item.get("officialProjectName")),
        clean(item.get("recordName")),
        *(CRIC_NAME_ALIASES.get(clean(item.get("officialProjectName")), [])),
    ]
    return find_summary_row(names, summary_rows)


def official_item_from_records(
    official_name: str,
    records: list[dict[str, Any]],
    data: dict[str, Any],
    summary_rows: dict[str, dict[str, Any]],
    plate_groups: dict[str, str],
) -> dict[str, Any]:
    first = records[0]
    permits = [clean(row.get("permit")) for row in records if clean(row.get("permit"))]
    issue_dates = [clean(row.get("issueDate")) for row in records if clean(row.get("issueDate"))]
    detail_urls = [clean(row.get("detailUrl")) for row in records if clean(row.get("detailUrl"))]
    residential_total = int(sum(number(row.get("approvedResidentialSuites")) for row in records))
    approved_total = int(sum(number(row.get("approvedTotalSuites")) for row in records))
    approved_residential_area = rounded(sum(number(row.get("approvedResidentialArea")) for row in records), 2)
    approved_sale_area = rounded(sum(number(row.get("approvedSaleArea")) for row in records), 2)
    item = {
        "officialProjectName": official_name,
        "district": clean(first.get("district")),
        "address": clean(first.get("location")),
        "developer": clean(first.get("developer")),
        "permits": permits,
        "issueDates": issue_dates,
        "residentialTotal": residential_total,
        "approvedTotalSuites": approved_total,
        "approvedResidentialArea": approved_residential_area,
        "approvedSaleArea": approved_sale_area,
        "detailUrls": detail_urls,
        "inventoryNote": (
            "住建委预售证详情页地上住宅批准套数汇总；车位、地下库房、商办证不计入住宅总套数；"
            f"不代表当前剩余套数。 计住宅证：{' / '.join(permits)}共{residential_total}套。"
        ),
        "source": "北京市住建委新建商品房项目公示",
    }
    cric = cric_summary_for_official({**item, "recordName": clean(first.get("recordName"))}, summary_rows)
    enrichment = OFFICIAL_PROJECT_ENRICHMENT.get(official_name, {})
    plate = enrichment.get("plate") or (cric or {}).get("plate") or clean(first.get("district"))
    item["plate"] = plate
    item["group"] = enrichment.get("group") or plate_groups.get(plate) or ""
    location = data.get("plateLocations", {}).get(plate, {})
    item["lat"] = enrichment.get("lat") or location.get("lat") or 0
    item["lng"] = enrichment.get("lng") or location.get("lng") or 0
    item["coordConfidence"] = enrichment.get("coordConfidence") or location.get("confidence") or "低"
    return item


def merge_official_new_launches(
    html: str,
    data: dict[str, Any],
    official_json: Path,
    summary_rows: dict[str, dict[str, Any]],
) -> tuple[str, list[dict[str, Any]]]:
    existing_match = OFFICIAL_RE.search(html)
    if not existing_match:
        raise RuntimeError("ZJW_OFFICIAL_NEW_LAUNCH_PROJECTS block not found")
    existing_items = json.loads(existing_match.group(1), strict=False)
    raw_records = [record for record in json.loads(official_json.read_text(encoding="utf-8")) if record.get("hasResidential")]
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in raw_records:
        grouped[clean(record.get("officialProjectName"))].append(record)
    plate_groups = plate_group_lookup(data)
    incoming = [
        official_item_from_records(name, records, data, summary_rows, plate_groups)
        for name, records in grouped.items()
    ]

    merged: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    additions: list[dict[str, Any]] = []
    for item in [*existing_items, *incoming]:
        name = clean(item.get("officialProjectName"))
        if not name:
            continue
        target = merged.get(name)
        if not target:
            target = dict(item)
            target["permits"] = []
            target["issueDates"] = []
            target["detailUrls"] = []
            counted_permits = set()
            merged[name] = target
            order.append(name)
        else:
            counted_permits = set(target.pop("_countedPermits", target.get("permits") or []))
        old_permits = set(target.get("permits") or [])
        for permit in item.get("permits") or []:
            if permit and permit not in target["permits"]:
                target["permits"].append(permit)
        for date in item.get("issueDates") or []:
            if date and date not in target["issueDates"]:
                target["issueDates"].append(date)
        for url in item.get("detailUrls") or []:
            if url and url not in target["detailUrls"]:
                target["detailUrls"].append(url)
        if item in incoming:
            new_permits = [permit for permit in item.get("permits", []) if permit and permit not in old_permits]
            if new_permits:
                additions.append({
                    "officialProjectName": name,
                    "permits": new_permits,
                    "issueDates": item.get("issueDates", []),
                    "residentialTotal": item.get("residentialTotal"),
                    "plate": item.get("plate"),
                })
        for field in ("district", "group", "plate", "lat", "lng", "coordConfidence", "address", "developer", "source"):
            if item.get(field):
                target[field] = item[field]
        for field in ("approvedResidentialArea", "approvedSaleArea", "approvedTotalSuites", "residentialTotal"):
            if item.get(field) is None:
                continue
            item_permits = {permit for permit in (item.get("permits") or []) if permit}
            if item_permits - counted_permits:
                if old_permits:
                    target[field] = rounded(number(target.get(field)) + number(item.get(field)), 2)
                else:
                    target[field] = item[field]
            else:
                target[field] = max(number(target.get(field)), number(item.get(field)))
        if item.get("inventoryNote"):
            target["inventoryNote"] = item["inventoryNote"]
        target["_countedPermits"] = counted_permits | {permit for permit in (item.get("permits") or []) if permit}

    output = []
    for name in order:
        item = merged[name]
        item.pop("_countedPermits", None)
        output.append(item)
    replacement = (
        "const ZJW_OFFICIAL_NEW_LAUNCH_PROJECTS = "
        + json.dumps(output, ensure_ascii=False, separators=(",", ":"))
        + ";\nconst ZJW_NEW_LAUNCH_INVENTORY_STATUS_OVERRIDES"
    )
    html = OFFICIAL_RE.sub(lambda _: replacement, html, count=1)
    html = html.replace(
        "const ZJW_NEW_LAUNCH_INVENTORY_STATUS_OVERRIDES\nconst ZJW_NEW_LAUNCH_INVENTORY_STATUS_OVERRIDES =",
        "const ZJW_NEW_LAUNCH_INVENTORY_STATUS_OVERRIDES =",
    )
    return html, additions


def existing_presale_match(record: dict[str, Any], project: dict[str, Any], alias_groups: list[list[str]]) -> bool:
    record_name = base.normalize_name(record.get("name"))
    permit = clean(record.get("permit"))
    permit_text = clean(project.get("summaryPresalePermit"))
    if permit and permit in permit_text:
        return True
    project_names = [base.normalize_name(name) for name in project_candidate_names(project, alias_groups)]
    return bool(record_name and any(name == record_name or (len(name) >= 4 and len(record_name) >= 4 and (name in record_name or record_name in name)) for name in project_names))


def build_summary_launch_project(
    item: dict[str, Any],
    summary: dict[str, Any],
    data: dict[str, Any],
    index: int,
) -> dict[str, Any]:
    official_name = clean(item.get("officialProjectName"))
    display_name = OFFICIAL_PROJECT_ENRICHMENT.get(official_name, {}).get("displayName") or summary["sourceProject"] or official_name
    months = data.get("months", [])
    monthly = {month: {"suites": 0, "area": 0, "price": 0, "amount": 0} for month in months}
    monthly[MONTH] = {
        "suites": summary["suites"],
        "area": summary["area"],
        "price": summary["price"] or calc_price(summary["amount"], summary["area"]),
        "amount": summary["amount"],
    }
    first_permit = clean((item.get("permits") or [""])[0])
    permit_id = re.search(r"\d{4}\)?\d+", first_permit)
    return {
        "id": f"zjw-launch-{permit_id.group(0) if permit_id else base.normalize_name(official_name) or index}",
        "group": clean(item.get("group")),
        "plate": clean(item.get("plate")),
        "project": display_name,
        "landDate": "",
        "status": "新取证",
        "x": 50,
        "y": 50,
        "lat": number(item.get("lat")),
        "lng": number(item.get("lng")),
        "monthly": monthly,
        "suites34": summary["suites"],
        "area34": summary["area"],
        "amount34": summary["amount"],
        "price4": monthly[MONTH]["price"],
        "amountAll": summary["amount"],
        "suitesAll": summary["suites"],
        "district": clean(item.get("district")) or summary.get("district", ""),
        "address": clean(item.get("address")),
        "coordSource": "住建委坐落位置 + 业务板块近似定位",
        "coordSourceUrl": clean((item.get("detailUrls") or [""])[0]),
        "coordConfidence": clean(item.get("coordConfidence")) or "低",
        "coordSystem": "GCJ-02近似板块中心",
        "matchedName": display_name,
        "summaryRecordName": official_name,
        "summaryPresalePermit": " / ".join(clean(value) for value in (item.get("permits") or []) if clean(value)),
        "summaryDeveloper": clean(item.get("developer")) or summary.get("developer", ""),
        "cricProjectName": summary["sourceProject"],
        "officialProjectName": official_name,
        "officialResidentialTotal": int(number(item.get("residentialTotal"))),
        "officialUnsignedSuites": None,
        "officialAvailableSuites": None,
        "officialUnsignedBlueSuites": 0,
        "officialBookedSuites": None,
        "officialContractSignedSuites": None,
        "officialFilingSuites": None,
        "officialSignedStatsSuites": None,
        "officialSignedStatsArea": None,
        "officialSignedStatsAvgPrice": None,
        "officialSignedSuites": None,
        "officialDetailSignedSuites": None,
        "officialInventoryEvidenceUrl": "\n".join(clean(value) for value in (item.get("detailUrls") or []) if clean(value)),
        "officialInventoryFetchedAt": "",
        "officialInventoryMatchStatus": "住建委新发预售证项目；8月成交取克而瑞项目累计供求汇总",
        "officialInventoryTotalAuditNote": clean(item.get("inventoryNote")),
        "approvedTotalSuites": int(number(item.get("approvedTotalSuites"))) or None,
        "source": "zjwNewLaunch",
        "presaleIssueRecords": base.build_official_project(item, months).get("presaleIssueRecords", []),
        "presaleIssueDates": sorted({clean(value) for value in item.get("issueDates", []) if clean(value)}),
        "augDataSource": "克尔瑞汇总",
        "augMatchedName": display_name,
        "augCricProjectName": summary["sourceProject"],
    }


def promote_summary_launch_projects(
    data: dict[str, Any],
    html: str,
    summary_rows: dict[str, dict[str, Any]],
    alias_groups: list[list[str]],
) -> list[dict[str, Any]]:
    official_items = json.loads(OFFICIAL_RE.search(html).group(1), strict=False)
    launch_projects = data.setdefault("launchProjects", [])
    existing = base.all_projects(data)
    promoted = []
    for index, item in enumerate(official_items):
        summary = cric_summary_for_official(item, summary_rows)
        if not summary or not number(summary.get("suites")):
            continue
        records = [
            {"name": clean(item.get("officialProjectName")), "permit": permit, "date": item.get("issueDates", [""])[idx] if idx < len(item.get("issueDates", [])) else ""}
            for idx, permit in enumerate(item.get("permits") or [])
        ]
        if any(any(existing_presale_match(record, project, alias_groups) for record in records) for project in existing):
            continue
        project = build_summary_launch_project(item, summary, data, index)
        launch_projects.append(project)
        existing.append(project)
        promoted.append({
            "project": project["project"],
            "officialProjectName": project["officialProjectName"],
            "summarySuites": summary["suites"],
            "summaryProject": summary["sourceProject"],
        })
    return promoted


def update_html(html: str, data: dict[str, Any], suffix: str, max_date: str) -> str:
    html = DATA_RE.sub(
        lambda _: "const DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n" + suffix,
        html,
        count=1,
    )
    html = re.sub(
        r'(<div class="cutoff"><i>▣</i><span>数据截至：</span><b>).*?(</b></div>)',
        rf"\g<1>{max_date.replace('/', '-')}\g<2>",
        html,
        count=1,
    )
    html = re.sub(
        r"const periodOptions = .*?;",
        "const periodOptions = " + json.dumps([[month, month] for month in reversed(data["months"])], ensure_ascii=False) + ";",
        html,
        count=1,
    )
    html = re.sub(
        r"const recentLaunchMonths = \[[^\]]*\]",
        "const recentLaunchMonths = " + json.dumps(data["months"][-2:], ensure_ascii=False),
        html,
        count=1,
    )
    html = re.sub(
        r'<script src="transaction_details\.js\?v=[^"]*"></script>',
        '<script src="transaction_details.js?v=20260904-august"></script>',
        html,
        count=1,
    )
    html = html.replace(
        'function projectSummarySource(p) {\n'
        '  if (p?.source === "zjwNewLaunch") return "新取证来源：北京市住建委预售许可证项目公示；成交明细暂未进入克尔瑞样本表";',
        'function projectSummarySource(p) {\n'
        '  if (p?.source === "zjwNewLaunch" && p?.augDataSource === "克尔瑞汇总") return "新取证来源：北京市住建委预售许可证项目公示；8月成交取克尔瑞项目累计供求汇总，逐套明细暂未覆盖";\n'
        '  if (p?.source === "zjwNewLaunch") return "新取证来源：北京市住建委预售许可证项目公示；成交明细暂未进入克尔瑞样本表";',
    )
    return html


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--html", type=Path, default=Path("index.html"))
    parser.add_argument("--detail-js", type=Path, default=Path("transaction_details.js"))
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--summary", type=Path, required=True)
    parser.add_argument("--official-json", type=Path, required=True)
    parser.add_argument("--report", type=Path, default=Path("august_update_report.json"))
    args = parser.parse_args()

    html, data, data_suffix = load_dashboard(args.html)
    alias_groups = base.parse_alias_groups(html)
    grouped, stats = base.read_workbook(args.xlsx)
    if list(grouped) != [MONTH]:
        raise RuntimeError(f"期望只更新 {MONTH}，但源表包含: {', '.join(grouped)}")
    matches, unmatched, ambiguous = base.match_groups(grouped, stats, base.all_projects(data), alias_groups)
    if unmatched or ambiguous:
        raise RuntimeError("存在未匹配或多义项目，停止写入")

    summary_rows = load_cric_summary(args.summary)
    validation = summary_validation(grouped, stats, matches, summary_rows, alias_groups)
    if validation["unmatchedSummary"] or validation["diffs"]:
        raise RuntimeError("8月明细与CRIC汇总未闭合，停止写入")

    old_details = load_transaction_details(args.detail_js)
    payload = base.build_transaction_payload(grouped, matches, alias_groups, args.xlsx.name)
    month_result = replace_august_dashboard_month(data, payload, alias_groups)

    html, official_additions = merge_official_new_launches(html, data, args.official_json, summary_rows)
    promoted_launch_projects = promote_summary_launch_projects(data, html, summary_rows, alias_groups)
    refresh_dashboard_aggregates(data)

    details = copy.deepcopy(old_details)
    details.setdefault("months", {})[MONTH] = payload["months"][MONTH]
    details["source"] = "项目成交明细.xlsx（25年7月-26年7月） + 陈麒亦需求.xlsx（26年8月）"
    recompute_transaction_summary(details)

    max_date = max(
        row["date"]
        for project in payload["months"][MONTH]["projects"].values()
        for row in project.get("rows", [])
    )
    historical_result = refresh_historical_sales_from_august_delta(
        data, old_details, details, alias_groups, max_date
    )
    write_transaction_details(args.detail_js, details)
    args.html.write_text(update_html(html, data, data_suffix, max_date), encoding="utf-8")

    report = {
        "month": MONTH,
        "source": str(args.xlsx),
        "summarySource": str(args.summary),
        "officialSource": str(args.official_json),
        "detailStats": stats,
        "matchedProjects": len(matches),
        "updatedDashboardProjects": len(month_result["matched"]),
        "zeroedProjects": month_result["zeroed"],
        "summaryValidation": validation,
        "historicalUpdate": historical_result,
        "officialNewLaunchAdditions": official_additions,
        "promotedLaunchProjects": promoted_launch_projects,
        "dashboardMonthTotal": next(row for row in data["monthlyTotals"] if row["month"] == MONTH),
        "transactionDetailMonthSummary": payload["months"][MONTH]["summary"],
        "transactionDetailSummary": details["summary"],
        "maxTradeDate": max_date,
    }
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "month": MONTH,
        "rawRows": stats["rawRows"],
        "matchedProjects": len(matches),
        "updatedDashboardProjects": len(month_result["matched"]),
        "officialNewLaunchAdditions": len(official_additions),
        "promotedLaunchProjects": len(promoted_launch_projects),
        "dashboardMonthTotal": report["dashboardMonthTotal"],
        "transactionDetailMonthSummary": report["transactionDetailMonthSummary"],
        "maxTradeDate": max_date,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

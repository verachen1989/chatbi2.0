#!/usr/bin/env python3
"""Build dashboard transaction detail JSON from the joined CRIC detail workbook."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DETAIL_SHEET = "成交明细_关联项目月度"
SUMMARY_SHEET = "项目月度汇总_按明细聚合"
DEFAULT_SCOPE = "普通住宅/别墅，已排除车库/车位"


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).lower()
    replacements = {
        "雲": "云",
        "澐": "云",
        "樹": "树",
        "華": "华",
        "鳴": "鸣",
        "號": "号",
        "鄕": "乡",
        "叁": "三",
        "贰": "二",
        "壹": "一",
        "玖": "九",
        "·": "",
        "•": "",
        ".": "",
        "。": "",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return re.sub(r"[\s\-—_/\\()（）【】\[\]《》“”\"'：:；;，,]+", "", text)


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def int_or_float(value: float) -> int | float:
    if abs(value - round(value)) < 0.000001:
        return int(round(value))
    return round(value, 2)


def row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}


def add_project_aliases(alias_map: dict[str, str], primary_key: str, group: dict[str, Any]) -> None:
    aliases = [
        group.get("projectName"),
        group.get("rawProjectName"),
        group.get("cricProjectName"),
        group.get("matchedProjectName"),
        group.get("summaryRecordName"),
    ]
    for alias in aliases:
        key = normalize_name(alias)
        if key and key != primary_key:
            alias_map.setdefault(key, primary_key)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--source", default="")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    detail_ws = wb[DETAIL_SHEET]
    summary_ws = wb[SUMMARY_SHEET]

    detail_headers = [clean(cell.value) for cell in next(detail_ws.iter_rows(min_row=1, max_row=1))]
    summary_headers = [clean(cell.value) for cell in next(summary_ws.iter_rows(min_row=1, max_row=1))]

    by_month_project: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    row_count = 0
    skipped_rows = 0
    excluded_parking_rows = 0
    property_counts: dict[str, int] = defaultdict(int)

    for values in detail_ws.iter_rows(min_row=2, values_only=True):
        row = row_dict(detail_headers, values)
        if clean(row.get("关联状态")) != "已关联":
            skipped_rows += 1
            continue
        month = clean(row.get("明细月份"))
        project_name = clean(row.get("去化表项目")) or clean(row.get("项目名称")) or clean(row.get("标题项目名"))
        if not month or not project_name:
            skipped_rows += 1
            continue
        property_type = clean(row.get("物业类型"))
        property_counts[property_type or "未填写"] += 1
        if property_type and property_type not in {"普通住宅", "别墅"}:
            excluded_parking_rows += 1
            continue
        key = normalize_name(project_name)
        month_projects = by_month_project[month]
        group = month_projects.get(key)
        if not group:
            group = {
                "projectName": project_name,
                "rawProjectName": project_name,
                "cricProjectName": clean(row.get("项目名称")),
                "matchedProjectName": clean(row.get("去化表项目")),
                "plate": clean(row.get("板块")),
                "group": clean(row.get("组团")),
                "rows": [],
                "summary": {"suites": 0, "area": 0, "amountWan": 0, "avgPrice": 0},
            }
            month_projects[key] = group
        total_wan = number(row.get("成交总价(万元)")) or number(row.get("成交总价(元)")) / 10000
        detail = {
            "date": clean(row.get("成交时间")),
            "permit": clean(row.get("预售证号")),
            "sourceProject": clean(row.get("项目名称")),
            "building": clean(row.get("楼栋名称")),
            "unit": clean(row.get("单元号")),
            "room": clean(row.get("室号")),
            "propertyType": property_type,
            "layout": clean(row.get("房型")),
            "area": int_or_float(number(row.get("成交面积(㎡)"))),
            "unitPrice": int_or_float(number(row.get("成交单价(元/㎡)"))),
            "totalWan": int_or_float(total_wan),
        }
        group["rows"].append(detail)
        row_count += 1

    summary_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for values in summary_ws.iter_rows(min_row=2, values_only=True):
        row = row_dict(summary_headers, values)
        month = clean(row.get("明细月份"))
        project_name = clean(row.get("去化表项目")) or clean(row.get("明细项目名称"))
        if not month or not project_name:
            continue
        summary_lookup[(month, normalize_name(project_name))] = row

    months: dict[str, Any] = {}
    for month, month_projects in sorted(by_month_project.items()):
        projects: dict[str, Any] = {}
        aliases: dict[str, str] = {}
        month_rows = 0
        for key, group in sorted(month_projects.items(), key=lambda item: item[1]["projectName"]):
            rows = group["rows"]
            rows.sort(key=lambda item: item["date"], reverse=True)
            summary_row = summary_lookup.get((month, key), {})
            suites = len(rows)
            area = sum(number(row["area"]) for row in rows)
            amount = sum(number(row["totalWan"]) for row in rows)
            avg_price = amount * 10000 / area if area else 0
            group["summary"] = {
                "suites": int(round(suites)),
                "area": int_or_float(area),
                "amountWan": int_or_float(amount),
                "avgPrice": int(round(avg_price)) if avg_price else 0,
            }
            group["plate"] = group.get("plate") or clean(summary_row.get("板块"))
            group["group"] = group.get("group") or clean(summary_row.get("组团"))
            group["cricProjectName"] = group.get("cricProjectName") or clean(summary_row.get("明细项目名称"))
            group["matchedProjectName"] = group.get("matchedProjectName") or clean(summary_row.get("去化表项目"))
            month_rows += len(rows)
            projects[key] = group
            add_project_aliases(aliases, key, group)
        months[month] = {
            "projects": projects,
            "aliases": aliases,
            "summary": {
                "projects": len({id(group) for group in month_projects.values()}),
                "rows": month_rows,
            },
        }

    payload = {
        "source": args.source or args.xlsx.name,
        "sheet": DETAIL_SHEET,
        "scope": DEFAULT_SCOPE,
        "months": months,
        "summary": {
            "months": len(months),
            "projects": sum(month["summary"]["projects"] for month in months.values()),
            "rows": row_count,
            "skippedRows": skipped_rows,
            "excludedParkingRows": excluded_parking_rows,
            "propertyTypes": dict(sorted(property_counts.items())),
        },
    }
    args.output.write_text(
        "window.TRANSACTION_DETAILS = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
        + "window.MAY_TRANSACTION_DETAILS = window.TRANSACTION_DETAILS.months['26年5月'] ? Object.assign({source:window.TRANSACTION_DETAILS.source,sheet:window.TRANSACTION_DETAILS.sheet,scope:window.TRANSACTION_DETAILS.scope,month:'2026年5月'}, window.TRANSACTION_DETAILS.months['26年5月']) : {summary:{projects:0,rows:0},projects:{}};\n",
        encoding="utf-8",
    )
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Update dashboard DATA for 2026 Jan-Apr using CRIC residential monthly summaries."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATA_RE = re.compile(r"const DATA = (.*?);\nconst DEFAULT_PERIODS", re.S)
MONTH_FILE_RE = re.compile(r"26\.(\d{1,2})(?:仅住宅)?")
METRIC_MONTHS = ["26年1月", "26年2月", "26年3月", "26年4月"]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).lower()
    replacements = {
        "雲": "云",
        "澐": "云",
        "樹": "树",
        "華": "华",
        "鳴": "鸣",
        "號": "号",
        "鄕": "乡",
        "臺": "台",
        "橒": "云",
        "叁": "三",
        "贰": "二",
        "壹": "一",
        "玖": "九",
        "·": "",
        "•": "",
        ".": "",
        "。": "",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[\s\-—_/\\()（）【】\[\]《》“”\"'：:；;，,]+", "", text)


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def rounded(value: float, digits: int = 2) -> float | int:
    value = round(float(value), digits)
    if abs(value - round(value)) < 0.000001:
        return int(round(value))
    return value


def row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}


def month_from_file(path: Path) -> str:
    match = MONTH_FILE_RE.search(path.name)
    if not match:
        raise ValueError(f"Cannot parse month from file name: {path.name}")
    return f"26年{int(match.group(1))}月"


def find_header_row(rows: list[tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows):
        if "项目名称" in [clean(value) for value in row]:
            return index
    raise ValueError("Cannot find header row with 项目名称")


def load_cric_summary(summary_dir: Path) -> dict[tuple[str, str], dict[str, Any]]:
    records: dict[tuple[str, str], dict[str, Any]] = {}
    for path in sorted(summary_dir.glob("*.xlsx")):
        if path.name.startswith(("~", ".~")):
            continue
        month = month_from_file(path)
        if month not in METRIC_MONTHS:
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            header_index = find_header_row(rows)
            headers = [clean(value) for value in rows[header_index]]
            for values in rows[header_index + 1 :]:
                row = row_dict(headers, values)
                project = clean(row.get("项目名称"))
                if not project:
                    continue
                area = number(row.get("成交面积(㎡)"))
                suites = int(round(number(row.get("成交套数(套)"))))
                amount_wan = number(row.get("成交金额(元)")) / 10000
                avg_price = int(round(number(row.get("成交均价(元/㎡)")))) if number(row.get("成交均价(元/㎡)")) else 0
                records[(month, normalize_name(project))] = {
                    "sourceProject": project,
                    "suites": suites,
                    "area": rounded(area, 2),
                    "price": avg_price,
                    "amount": rounded(amount_wan, 4),
                }
        finally:
            wb.close()
    return records


def candidate_names(project: dict[str, Any]) -> list[str]:
    names: list[str] = []
    for field in ("cricProjectName", "matchedName", "project", "summaryRecordName", "officialProjectName"):
        value = clean(project.get(field))
        if not value:
            continue
        names.extend(re.split(r"[；;、/\n]+", value))
    output = []
    for name in names:
        name = clean(name)
        if name and name not in output:
            output.append(name)
    return output


def find_summary(summary: dict[tuple[str, str], dict[str, Any]], month: str, project: dict[str, Any]) -> tuple[dict[str, Any] | None, str]:
    for name in candidate_names(project):
        key = normalize_name(name)
        if key and (month, key) in summary:
            return summary[(month, key)], name
    return None, ""


def calc_price(amount: float, area: float) -> int:
    return int(round(amount * 10000 / area)) if area else 0


def recalc_project(project: dict[str, Any], months: list[str]) -> None:
    recent_months = months[-2:]
    project["suites34"] = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for month in recent_months)
    project["area34"] = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for month in recent_months)
    project["amount34"] = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for month in recent_months)
    latest = project.get("monthly", {}).get(recent_months[-1], {}) if recent_months else {}
    project["price4"] = latest.get("price", 0)
    project["suitesAll"] = sum(number(metric.get("suites")) for metric in project.get("monthly", {}).values())
    project["amountAll"] = sum(number(metric.get("amount")) for metric in project.get("monthly", {}).values())


def aggregate_month(projects: list[dict[str, Any]], month: str) -> dict[str, Any]:
    suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects)
    area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects)
    amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects)
    return {"month": month, "suites": rounded(suites), "area": rounded(area, 2), "amount": rounded(amount, 4), "price": calc_price(amount, area)}


def refresh_aggregates(data: dict[str, Any]) -> None:
    months = data["months"]
    recent_months = months[-2:]
    for collection_name in ("projects", "launchProjects"):
        for project in data.get(collection_name, []):
            recalc_project(project, months)

    projects = data["projects"]
    recent_suites = sum(number(project.get("monthly", {}).get(month, {}).get("suites")) for project in projects for month in recent_months)
    recent_area = sum(number(project.get("monthly", {}).get(month, {}).get("area")) for project in projects for month in recent_months)
    recent_amount = sum(number(project.get("monthly", {}).get(month, {}).get("amount")) for project in projects for month in recent_months)
    data["totals"] = {
        "projects": len(projects),
        "plates": len({project.get("plate") for project in projects}),
        "active": sum(project.get("status") == "在售" for project in projects),
        "suites34": rounded(recent_suites),
        "amount34": rounded(recent_amount, 4),
        "area34": rounded(recent_area, 2),
        "avgPrice34": calc_price(recent_amount, recent_area),
    }
    data["monthlyTotals"] = [aggregate_month(projects, month) for month in months]

    group_totals = []
    for group in dict.fromkeys(project.get("group") for project in projects):
        group_projects = [project for project in projects if project.get("group") == group]
        group_totals.append(
            {
                "group": group,
                "projects": len(group_projects),
                "suites34": rounded(sum(number(project.get("suites34")) for project in group_projects)),
                "amount34": rounded(sum(number(project.get("amount34")) for project in group_projects), 4),
                "color": data.get("colors", {}).get(group, "#8ea0b8"),
            }
        )
    data["groupTotals"] = group_totals
    data["colors"] = {row["group"]: row["color"] for row in group_totals}
    policy = data.setdefault("sourcePolicy", {})
    policy.pop("2026年1-5月成交明细", None)
    policy["2026年1-4月成交汇总"] = "克尔瑞项目累计供求汇总（仅住宅）"
    policy["2026年1-4月成交明细"] = "克尔瑞交易明细（普通住宅/别墅）"
    policy["26年5月成交明细"] = "克尔瑞交易明细（普通住宅/别墅）"
    policy["2026年5月"] = "保持原页面数据，本次未更新"


def update_collection(collection: list[dict[str, Any]], summary: dict[tuple[str, str], dict[str, Any]]) -> dict[str, int]:
    stats = defaultdict(int)
    for project in collection:
        for month in METRIC_MONTHS:
            row, matched_name = find_summary(summary, month, project)
            if not row:
                stats[f"{month}_未命中"] += 1
                continue
            project.setdefault("monthly", {})[month] = {
                "suites": row["suites"],
                "area": row["area"],
                "price": row["price"],
                "amount": row["amount"],
            }
            project["janAprDataSource"] = "克尔瑞"
            project["janAprMatchedName"] = matched_name
            stats[f"{month}_已更新"] += 1
    return dict(stats)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--html", type=Path, default=Path("index.html"))
    parser.add_argument("--summary-dir", type=Path, required=True)
    args = parser.parse_args()

    html = args.html.read_text(encoding="utf-8")
    match = DATA_RE.search(html)
    if not match:
        raise RuntimeError("DATA block not found")
    data = json.loads(match.group(1), strict=False)
    summary = load_cric_summary(args.summary_dir)
    stats = {"projects": update_collection(data["projects"], summary)}
    if data.get("launchProjects"):
        stats["launchProjects"] = update_collection(data["launchProjects"], summary)
    refresh_aggregates(data)
    replacement = "const DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\nconst DEFAULT_PERIODS"
    args.html.write_text(DATA_RE.sub(lambda _: replacement, html, count=1), encoding="utf-8")
    print(json.dumps({"summaryRows": len(summary), "stats": stats}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
